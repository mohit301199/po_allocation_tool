import streamlit as st
import pandas as pd
from sqlalchemy import create_engine, text
import bcrypt
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment

from barcode import Code128
from barcode.writer import ImageWriter

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak

import tempfile
import os

from marketing_dashboard import show_marketing_dashboard


# =====================================================
# PAGE CONFIG
# =====================================================

st.set_page_config(
    page_title="PO Allocation",
    page_icon="package",
    layout="wide"
)


# =====================================================
# CSS
# =====================================================

st.markdown("""
<style>
.stApp {
    background: linear-gradient(135deg,#f8fafc,#eef2ff,#fff7ed);
    color: #0f172a;
}

.page-header {
    background: rgba(255,255,255,0.86);
    border: 1px solid rgba(148,163,184,0.35);
    border-left: 5px solid #2563eb;
    padding: 18px 22px;
    border-radius: 14px;
    margin-bottom: 22px;
    box-shadow: 0 10px 28px rgba(15,23,42,0.06);
}

.page-title {
    font-size: 28px;
    line-height: 1.2;
    font-weight: 900;
    color: #0f172a;
}

.page-subtitle {
    font-size: 14px;
    color: #475569;
    margin-top: 5px;
    font-weight: 600;
}

[data-testid="stSidebar"] {
    background: linear-gradient(180deg,#0f172a,#1e1b4b,#312e81);
}

[data-testid="stSidebar"] * {
    color: white !important;
}

div.stButton > button {
    background: linear-gradient(135deg,#2563eb,#7c3aed);
    color: white !important;
    border: none;
    border-radius: 14px;
    font-weight: 800;
    padding: 12px 22px;
    font-size: 15px;
}

div.stDownloadButton > button {
    background: linear-gradient(135deg,#2563eb,#7c3aed) !important;
    color: white !important;
    border: none !important;
    border-radius: 14px !important;
    font-weight: 800 !important;
    padding: 12px 22px !important;
    font-size: 15px !important;
}

div.stDownloadButton > button:hover {
    background: linear-gradient(135deg,#1d4ed8,#6d28d9) !important;
    color: white !important;
}

div[data-testid="stFileUploader"] {
    background: rgba(255,255,255,0.9);
    border: 2px dashed #2563eb;
    border-radius: 18px;
    padding: 18px;
}

div[data-testid="stFileUploader"] button {
    background: #2563eb !important;
    color: white !important;
    font-weight: 800 !important;
}

div[data-testid="stFileUploader"] button * {
    color: white !important;
}

label,
div[data-testid="stWidgetLabel"] p {
    color: #0f172a !important;
    font-weight: 800 !important;
    font-size: 16px !important;
}

.metric-card {
    padding: 24px;
    border-radius: 22px;
    color: white;
    text-align: center;
}

.blue { background: linear-gradient(135deg,#2563eb,#7c3aed); }
.green { background: linear-gradient(135deg,#059669,#22c55e); }
.orange { background: linear-gradient(135deg,#f97316,#f59e0b); }
.red { background: linear-gradient(135deg,#dc2626,#f43f5e); }

.metric-title {
    font-size: 15px;
    font-weight: 700;
}

.metric-value {
    font-size: 34px;
    font-weight: 900;
}

.success-box {
    background: #dcfce7;
    border-left: 6px solid #22c55e;
    padding: 15px;
    border-radius: 14px;
    color: #166534;
    font-weight: 800;
}
</style>
""", unsafe_allow_html=True)


def render_page_header(title, subtitle):
    st.markdown(f"""
    <div class="page-header">
        <div class="page-title">{title}</div>
        <div class="page-subtitle">{subtitle}</div>
    </div>
    """, unsafe_allow_html=True)


# =====================================================
# DATABASE
# =====================================================

@st.cache_resource
def get_engine():
    if "DATABASE_URL" not in st.secrets:
        st.error(
            "DATABASE_URL is missing. Please add your Supabase connection string "
            "in .streamlit/secrets.toml."
        )
        st.stop()

    return create_engine(
        st.secrets["DATABASE_URL"],
        pool_pre_ping=True,
        pool_recycle=1800,
        pool_size=5,
        max_overflow=5,
    )


engine = get_engine()


def make_params_key(params):
    if not params:
        return tuple()

    return tuple(sorted(params.items()))


@st.cache_data(ttl=20, show_spinner=False)
def db_read_cached(query, params_key):
    params = dict(params_key)

    with engine.connect() as connection:
        return pd.read_sql(text(query), connection, params=params)


def db_read(query, params=None, use_cache=True):
    params_key = make_params_key(params)

    if use_cache:
        return db_read_cached(query, params_key).copy()

    with engine.connect() as connection:
        return pd.read_sql(text(query), connection, params=params or {})


def db_execute(query, params=None, clear_cache=True):
    with engine.begin() as connection:
        connection.execute(text(query), params or {})

    if clear_cache:
        st.cache_data.clear()


def db_execute_many(query, params_list, clear_cache=True):
    if not params_list:
        return

    with engine.begin() as connection:
        connection.execute(text(query), params_list)

    if clear_cache:
        st.cache_data.clear()


def ensure_performance_indexes():
    index_queries = [
        """
        ALTER TABLE allocation_tracker
        ADD COLUMN IF NOT EXISTS appointment_no TEXT
        """,
        """
        ALTER TABLE allocation_tracker
        ADD COLUMN IF NOT EXISTS appointment_date TEXT
        """,
        """
        ALTER TABLE allocation_tracker
        ADD COLUMN IF NOT EXISTS fcn TEXT
        """,
        """
        ALTER TABLE allocation_tracker
        ADD COLUMN IF NOT EXISTS order_id TEXT
        """,
        """
        ALTER TABLE allocation_tracker
        ADD COLUMN IF NOT EXISTS buyer_code TEXT
        """,
        """
        ALTER TABLE allocation_tracker
        ADD COLUMN IF NOT EXISTS pending_amount REAL DEFAULT 0
        """,
        """
        ALTER TABLE allocation_tracker
        ADD COLUMN IF NOT EXISTS billed_qty REAL DEFAULT 0
        """,
        """
        ALTER TABLE allocation_tracker
        ADD COLUMN IF NOT EXISTS balance_to_bill REAL DEFAULT 0
        """,
        """
        ALTER TABLE allocation_tracker
        ADD COLUMN IF NOT EXISTS billing_source TEXT
        """,
        """
        CREATE INDEX IF NOT EXISTS idx_allocation_tracker_sent_billing
        ON allocation_tracker (sent_for_billing, billing_done)
        """,
        """
        CREATE INDEX IF NOT EXISTS idx_allocation_tracker_fsn_rr_sap
        ON allocation_tracker (fsn, rr_warehouse, sap_code)
        """,
        """
        CREATE INDEX IF NOT EXISTS idx_allocation_tracker_po_fsn_rr_fk
        ON allocation_tracker (po_no, fsn, rr_warehouse, fk_warehouse)
        """,
        """
        CREATE INDEX IF NOT EXISTS idx_allocation_tracker_sales_match
        ON allocation_tracker (po_no, fsn, sap_code, rr_warehouse)
        """,
        """
        CREATE INDEX IF NOT EXISTS idx_allocation_tracker_invoice
        ON allocation_tracker (invoice_no)
        """,
        """
        CREATE INDEX IF NOT EXISTS idx_allocation_tracker_allocation_date
        ON allocation_tracker (allocation_date)
        """,
        """
        CREATE INDEX IF NOT EXISTS idx_allocation_tracker_billing_date
        ON allocation_tracker (billing_date)
        """,
        """
        CREATE INDEX IF NOT EXISTS idx_allocation_tracker_sent_date
        ON allocation_tracker (sent_date)
        """,
        """
        CREATE INDEX IF NOT EXISTS idx_allocation_tracker_order_id
        ON allocation_tracker (order_id)
        """,
        """
        CREATE INDEX IF NOT EXISTS idx_allocation_tracker_id
        ON allocation_tracker (id)
        """,
        """
        CREATE TABLE IF NOT EXISTS appointment_tracker (
            id SERIAL PRIMARY KEY,
            upload_date TEXT,
            appointment_no TEXT,
            appointment_date TEXT,
            po_no TEXT,
            fsn TEXT,
            rr_warehouse TEXT,
            fk_warehouse TEXT,
            appointment_qty REAL,
            remark TEXT
        )
        """,
        """
        CREATE INDEX IF NOT EXISTS idx_appointment_tracker_match
        ON appointment_tracker (fsn, fk_warehouse, po_no, appointment_date)
        """,
    ]

    for query in index_queries:
        db_execute(query, clear_cache=False)


if "performance_indexes_checked" not in st.session_state:
    try:
        ensure_performance_indexes()
    except Exception:
        pass
    st.session_state.performance_indexes_checked = True


# =====================================================
# AUTHENTICATION
# =====================================================

def hash_password(password):
    return bcrypt.hashpw(
        password.encode("utf-8"),
        bcrypt.gensalt()
    ).decode("utf-8")


def verify_password(password, password_hash):
    try:
        return bcrypt.checkpw(
            password.encode("utf-8"),
            password_hash.encode("utf-8")
        )
    except Exception:
        return False


def auth_clean_text(x):
    if x is None:
        return ""
    return str(x).replace("\xa0", " ").strip()


def get_user_count():
    result = db_read("SELECT COUNT(*) AS user_count FROM app_users")
    return int(result.iloc[0]["user_count"])


def log_activity(action, details=""):
    username = st.session_state.get("username", "")

    if not username:
        return

    db_execute("""
        INSERT INTO activity_log (username, action, details)
        VALUES (:username, :action, :details)
    """, {
        "username": username,
        "action": action,
        "details": details
    }, clear_cache=False)


def create_first_admin_screen():
    render_page_header(
        "Create Admin User",
        "Set up the first login for this cloud tracker."
    )

    st.info("No app users exist yet. Create the first Admin account to continue.")

    username = st.text_input("Admin Username")
    password = st.text_input("Admin Password", type="password")
    confirm_password = st.text_input("Confirm Password", type="password")

    if st.button("Create Admin"):
        username = auth_clean_text(username)

        if username == "":
            st.error("Please enter an admin username.")

        elif len(password) < 6:
            st.error("Password should be at least 6 characters.")

        elif password != confirm_password:
            st.error("Passwords do not match.")

        else:
            db_execute("""
                INSERT INTO app_users (username, password_hash, role, active)
                VALUES (:username, :password_hash, :role, TRUE)
            """, {
                "username": username,
                "password_hash": hash_password(password),
                "role": "Admin"
            })

            st.success("Admin user created. Please log in.")
            st.rerun()


def login_screen():
    render_page_header(
        "Login",
        "Sign in to access the PO allocation tracker."
    )

    username = st.text_input("Username")
    password = st.text_input("Password", type="password")

    if st.button("Login"):
        username = auth_clean_text(username)

        users = db_read("""
            SELECT username, password_hash, role
            FROM app_users
            WHERE username = :username
            AND active = TRUE
        """, {"username": username})

        if users.empty:
            st.error("Invalid username or password.")
            return

        user = users.iloc[0]

        if not verify_password(password, user["password_hash"]):
            st.error("Invalid username or password.")
            return

        st.session_state.logged_in = True
        st.session_state.username = user["username"]
        st.session_state.role = user["role"]
        log_activity("login", "User logged in")
        st.rerun()


def get_allowed_screens(role):
    if role == "Admin":
        return [
            "Dashboard Summary",
            "Marketing Dashboard",
            "Upload & Allocate",
            "Allocation Tracker",
            "Billing Summary",
            "Open Allocation Qty",
            "User Management"
        ]

    if role == "Ops":
        return [
            "Dashboard Summary",
            "Marketing Dashboard",
            "Upload & Allocate",
            "Allocation Tracker",
            "Billing Summary",
            "Open Allocation Qty"
        ]

    if role == "Billing":
        return [
            "Dashboard Summary",
            "Marketing Dashboard",
            "Allocation Tracker",
            "Billing Summary",
            "Open Allocation Qty"
        ]

    return [
        "Dashboard Summary",
        "Marketing Dashboard",
        "Billing Summary",
        "Open Allocation Qty"
    ]


if "logged_in" not in st.session_state:
    st.session_state.logged_in = False


if get_user_count() == 0:
    create_first_admin_screen()
    st.stop()


if not st.session_state.logged_in:
    login_screen()
    st.stop()


# =====================================================
# SESSION STATE
# =====================================================

if "stock_df" not in st.session_state:
    st.session_state.stock_df = None

if "pending_df" not in st.session_state:
    st.session_state.pending_df = None

if "allocation_df" not in st.session_state:
    st.session_state.allocation_df = None


# =====================================================
# HELPER FUNCTIONS
# =====================================================

def clean_text(x):
    if pd.isna(x):
        return ""
    return str(x).replace("\xa0", " ").strip()


def clean_number(x):
    if pd.isna(x):
        return 0
    try:
        return float(str(x).replace(",", "").strip())
    except:
        return 0


def normalize_header(col):
    col = str(col)
    col = col.replace("\xa0", " ")
    col = col.replace("\n", " ")
    col = col.replace("\r", " ")
    col = col.replace("\t", " ")
    col = " ".join(col.split())
    return col.strip()


def normalize_columns(df):
    df = df.copy()
    df.columns = [normalize_header(c) for c in df.columns]
    return df


ALLOWED_STOCK_SITES = [
    "WBKOL02",
    "KABNG01",
    "TSHYD01",
    "HRAMB01",
    "MHBHW02",
    "PBTEP01",
    "HRAMB02",
    "HRAMB03",
    "MHBHW01",
    "KABNG02",
]

ALLOWED_STOCK_LOCATIONS = ["FGI", "ECOM"]


def first_existing_column(df, options):
    for option in options:
        if option in df.columns:
            return option
    return None


def standardize_stock_file(stock_df, show_messages=False):
    stock_df = normalize_columns(stock_df)

    old_format_cols = ["FSN", "RR Warehouse", "SAP Code", "Stock"]

    if all(col in stock_df.columns for col in old_format_cols):
        return stock_df

    item_code_col = first_existing_column(stock_df, ["Item Code", "Item code", "ItemCode"])
    available_qty_col = first_existing_column(stock_df, ["Available Qty", "Available Qty.", "AvailableQty"])
    site_col = first_existing_column(stock_df, ["Site", "SITE"])
    custom_location_col = first_existing_column(
        stock_df,
        ["Custom Location", "CustomLocation", "CUSTOM LOCATION", "CUSTOMLOCATION"]
    )
    fsn_col = first_existing_column(stock_df, ["FSN", "fsn"])

    system_stock_cols = [
        item_code_col,
        available_qty_col,
        site_col,
        custom_location_col,
        fsn_col,
    ]

    if not all(system_stock_cols):
        return stock_df

    original_rows = len(stock_df)

    stock_df = stock_df.copy()
    stock_df[item_code_col] = stock_df[item_code_col].apply(clean_text)
    stock_df[site_col] = stock_df[site_col].apply(clean_text).str.upper()
    stock_df[custom_location_col] = stock_df[custom_location_col].apply(clean_text).str.upper()
    stock_df[fsn_col] = stock_df[fsn_col].apply(clean_text)
    stock_df[available_qty_col] = stock_df[available_qty_col].apply(clean_number)

    stock_df = stock_df[
        stock_df[custom_location_col].isin(ALLOWED_STOCK_LOCATIONS) &
        stock_df[site_col].isin(ALLOWED_STOCK_SITES) &
        (stock_df[available_qty_col] > 0) &
        (stock_df[fsn_col] != "")
    ]

    standardized_df = stock_df.rename(columns={
        fsn_col: "FSN",
        site_col: "RR Warehouse",
        item_code_col: "SAP Code",
        available_qty_col: "Stock",
    })[["FSN", "RR Warehouse", "SAP Code", "Stock"]]

    standardized_df = standardized_df.groupby(
        ["FSN", "RR Warehouse", "SAP Code"],
        as_index=False
    )["Stock"].sum()

    if show_messages:
        st.info(
            "System stock file detected. "
            f"Filtered {original_rows:,} rows to {len(standardized_df):,} usable stock rows "
            "using CustomLocation FGI/ECOM and allowed SITE list."
        )

    return standardized_df


def standardize_pending_file(pending_df, show_messages=False):
    pending_df = normalize_columns(pending_df)

    old_format_cols = ["PO No.", "FSN", "Title", "RR Warehouse", "FK Warehouse", "Pending Qty."]

    if all(col in pending_df.columns for col in old_format_cols):
        if "Order ID" not in pending_df.columns:
            pending_df["Order ID"] = ""
        if "Buyer Code" not in pending_df.columns:
            pending_df["Buyer Code"] = ""
        if "Pending Amount" not in pending_df.columns:
            pending_df["Pending Amount"] = 0
        return pending_df

    po_col = first_existing_column(pending_df, ["PONo", "PO No.", "PO No", "PO Number", "PO"])
    order_id_col = first_existing_column(pending_df, ["Order Id", "Order ID", "OrderId"])
    buyer_code_col = first_existing_column(pending_df, ["Buyer Code", "BuyerCode"])
    fsn_col = first_existing_column(pending_df, ["FSN", "fsn"])
    title_col = first_existing_column(pending_df, ["Item Description", "Title", "Product Title"])
    rr_col = first_existing_column(pending_df, ["Site Id", "Site ID", "SITE", "Site"])
    fk_col = first_existing_column(pending_df, ["FK FC", "FK Warehouse", "FK WH"])
    pending_qty_col = first_existing_column(pending_df, ["Pending Qty", "Pending Qty.", "Pending Quantity"])
    pending_amount_col = first_existing_column(pending_df, ["Pending Amt", "Pending Amount", "Pending Amt."])
    status_col = first_existing_column(pending_df, ["Status", "Order Status"])

    system_pending_cols = [
        po_col,
        order_id_col,
        buyer_code_col,
        fsn_col,
        title_col,
        rr_col,
        fk_col,
        pending_qty_col,
        pending_amount_col,
        status_col,
    ]

    if not all(system_pending_cols):
        return pending_df

    original_rows = len(pending_df)

    pending_df = pending_df.copy()
    pending_df[status_col] = pending_df[status_col].apply(clean_text)
    pending_df = pending_df[
        pending_df[status_col].str.lower().isin(["hold", "pending"])
    ]

    for col in [po_col, order_id_col, buyer_code_col, fsn_col, title_col, rr_col, fk_col]:
        pending_df[col] = pending_df[col].apply(clean_text)

    pending_df[rr_col] = pending_df[rr_col].str.upper()
    pending_df[pending_qty_col] = pending_df[pending_qty_col].apply(clean_number)
    pending_df[pending_amount_col] = pending_df[pending_amount_col].apply(clean_number)

    blank_fsn_rows = len(pending_df[pending_df[fsn_col] == ""])

    pending_df = pending_df[
        (pending_df[fsn_col] != "") &
        (pending_df[pending_qty_col] > 0)
    ]

    standardized_df = pending_df.rename(columns={
        po_col: "PO No.",
        order_id_col: "Order ID",
        buyer_code_col: "Buyer Code",
        fsn_col: "FSN",
        title_col: "Title",
        rr_col: "RR Warehouse",
        fk_col: "FK Warehouse",
        pending_qty_col: "Pending Qty.",
        pending_amount_col: "Pending Amount",
    })[
        [
            "PO No.",
            "Order ID",
            "Buyer Code",
            "FSN",
            "Title",
            "RR Warehouse",
            "FK Warehouse",
            "Pending Qty.",
            "Pending Amount",
        ]
    ]

    standardized_df = standardized_df.groupby(
        ["PO No.", "Order ID", "Buyer Code", "FSN", "Title", "RR Warehouse", "FK Warehouse"],
        as_index=False
    ).agg({
        "Pending Qty.": "sum",
        "Pending Amount": "sum",
    })

    if show_messages:
        st.info(
            "System pending file detected. "
            f"Included Hold/Pending rows and converted {original_rows:,} source rows "
            f"to {len(standardized_df):,} usable pending rows. "
            f"Skipped {blank_fsn_rows:,} rows with blank FSN."
        )

    return standardized_df


def get_allocation_tracker_columns():
    columns_df = db_read("""
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = 'public'
        AND table_name = 'allocation_tracker'
    """)

    return set(columns_df["column_name"].tolist())


def get_tracker_df():
    return db_read("SELECT * FROM allocation_tracker ORDER BY id DESC")


def get_open_allocation_qty():
    tracker_columns = get_allocation_tracker_columns()

    if "billed_qty" in tracker_columns:
        query = """
        SELECT
            fsn,
            rr_warehouse,
            sap_code,
            SUM(GREATEST(allocated_qty - COALESCE(billed_qty, 0), 0)) AS open_alloc_qty
        FROM allocation_tracker
        WHERE sent_for_billing='Yes'
        AND GREATEST(allocated_qty - COALESCE(billed_qty, 0), 0) > 0
        GROUP BY fsn, rr_warehouse, sap_code
        """

    else:
        query = """
        SELECT
            fsn,
            rr_warehouse,
            sap_code,
            SUM(allocated_qty) AS open_alloc_qty
        FROM allocation_tracker
        WHERE sent_for_billing='Yes'
        AND (billing_done IS NULL OR billing_done!='Yes')
        GROUP BY fsn, rr_warehouse, sap_code
        """

    return db_read(query)


def get_existing_po_allocation_qty():
    query = """
    SELECT
        po_no,
        fsn,
        rr_warehouse,
        fk_warehouse,
        SUM(allocated_qty) AS already_allocated_qty
    FROM allocation_tracker
    GROUP BY po_no, fsn, rr_warehouse, fk_warehouse
    """

    return db_read(query)


def get_appointment_df():
    try:
        appointment_df = db_read("""
        SELECT
            a.*,
            COALESCE(used.used_qty, 0) AS used_qty,
            GREATEST(a.appointment_qty - COALESCE(used.used_qty, 0), 0) AS appointment_balance_qty
        FROM appointment_tracker a
        LEFT JOIN (
            SELECT
                appointment_no,
                fsn,
                fk_warehouse,
                po_no,
                SUM(allocated_qty) AS used_qty
            FROM allocation_tracker
            WHERE appointment_no IS NOT NULL
            AND TRIM(appointment_no) <> ''
            GROUP BY appointment_no, fsn, fk_warehouse, po_no
        ) used
        ON COALESCE(a.appointment_no, '') = COALESCE(used.appointment_no, '')
        AND COALESCE(a.fsn, '') = COALESCE(used.fsn, '')
        AND COALESCE(a.fk_warehouse, '') = COALESCE(used.fk_warehouse, '')
        AND COALESCE(a.po_no, '') = COALESCE(used.po_no, '')
        WHERE GREATEST(a.appointment_qty - COALESCE(used.used_qty, 0), 0) > 0
        ORDER BY a.appointment_date, a.id
        """)
    except Exception:
        appointment_df = pd.DataFrame()

    return appointment_df


def find_appointment_matches(appointment_df, po, fsn, rr, fk):
    if appointment_df.empty:
        return pd.DataFrame()

    appt_df = appointment_df.copy()

    for col in ["po_no", "fsn", "rr_warehouse", "fk_warehouse", "appointment_no"]:
        if col not in appt_df.columns:
            appt_df[col] = ""
        appt_df[col] = appt_df[col].apply(clean_text)

    mask = (
        (appt_df["fsn"] == fsn) &
        (appt_df["fk_warehouse"] == fk) &
        ((appt_df["po_no"] == "") | (appt_df["po_no"] == po)) &
        (appt_df["appointment_balance_qty"].apply(clean_number) > 0)
    )

    return appt_df[mask].copy()


# =========================
# BILLING SUMMARY
# =========================

def get_billing_summary(tracker_ids=None):
    tracker_columns = get_allocation_tracker_columns()
    qty_expr = "COALESCE(billed_qty, allocated_qty)" if "billed_qty" in tracker_columns else "allocated_qty"
    id_filter = ""

    if tracker_ids:
        clean_ids = [int(clean_number(tracker_id)) for tracker_id in tracker_ids if clean_number(tracker_id) > 0]

        if clean_ids:
            id_filter = f"AND id IN ({', '.join(map(str, clean_ids))})"

    query = f"""
    SELECT
        id,
        invoice_no,
        po_no,
        fsn,
        title,
        rr_warehouse,
        fk_warehouse,
        sap_code,
        {qty_expr} AS allocated_qty
    FROM allocation_tracker
    WHERE sent_for_billing = 'Yes'
    AND invoice_no IS NOT NULL
    AND TRIM(invoice_no) <> ''
    {id_filter}
    ORDER BY invoice_no, po_no, fsn, sap_code
    """

    return db_read(query)


def get_sent_for_billing_download_df(tracker_df):
    if tracker_df.empty or "sent_for_billing" not in tracker_df.columns:
        return pd.DataFrame()

    export_df = tracker_df[
        tracker_df["sent_for_billing"].fillna("").str.strip().str.lower() == "yes"
    ].copy()

    if export_df.empty:
        return pd.DataFrame()

    export_columns = [
        "id",
        "sent_date",
        "po_no",
        "order_id",
        "buyer_code",
        "fsn",
        "title",
        "rr_warehouse",
        "fk_warehouse",
        "sap_code",
        "fcn",
        "appointment_no",
        "appointment_date",
        "pending_amount",
        "allocated_qty",
        "billed_qty",
        "balance_to_bill",
        "invoice_no",
        "billing_date",
        "billing_done",
        "remark",
    ]

    export_columns = [col for col in export_columns if col in export_df.columns]
    export_df = export_df[export_columns]

    return export_df.rename(columns={
        "id": "Tracker ID",
        "sent_date": "Sent Date",
        "po_no": "PO No.",
        "order_id": "Order ID",
        "buyer_code": "Buyer Code",
        "fsn": "FSN",
        "title": "Title",
        "rr_warehouse": "RR Warehouse",
        "fk_warehouse": "FK Warehouse",
        "sap_code": "SAP Code",
        "fcn": "FCN",
        "appointment_no": "Appointment No.",
        "appointment_date": "Appointment Date",
        "pending_amount": "Pending Amount",
        "allocated_qty": "Allocated Qty.",
        "billed_qty": "Billed Qty.",
        "balance_to_bill": "Balance To Bill",
        "invoice_no": "Invoice No.",
        "billing_date": "Billing Date",
        "billing_done": "Billing Done",
        "remark": "Remark",
    })


def get_billing_update_template_df(tracker_df):
    if tracker_df.empty:
        return pd.DataFrame()

    template_df = tracker_df.copy()

    export_columns = [
        "id",
        "po_no",
        "order_id",
        "buyer_code",
        "fsn",
        "title",
        "rr_warehouse",
        "fk_warehouse",
        "sap_code",
        "fcn",
        "appointment_no",
        "appointment_date",
        "allocated_qty",
        "sent_for_billing",
        "sent_date",
        "billed_qty",
        "balance_to_bill",
        "invoice_no",
        "billing_date",
        "billing_done",
        "remark",
    ]

    export_columns = [col for col in export_columns if col in template_df.columns]
    template_df = template_df[export_columns]

    template_df = template_df.rename(columns={
        "id": "Tracker ID",
        "po_no": "PO No.",
        "order_id": "Order ID",
        "buyer_code": "Buyer Code",
        "fsn": "FSN",
        "title": "Title",
        "rr_warehouse": "RR Warehouse",
        "fk_warehouse": "FK Warehouse",
        "sap_code": "SAP Code",
        "fcn": "FCN",
        "appointment_no": "Appointment No.",
        "appointment_date": "Appointment Date",
        "allocated_qty": "Allocated Qty.",
        "sent_for_billing": "Sent For Billing (Yes/No)",
        "sent_date": "Sent Date (YYYY-MM-DD)",
        "billed_qty": "Billed Qty.",
        "balance_to_bill": "Balance To Bill",
        "invoice_no": "Invoice No.",
        "billing_date": "Billing Date (YYYY-MM-DD)",
        "billing_done": "Billing Done (Yes/No)",
        "remark": "Remark",
    })

    for col in ["Sent Date (YYYY-MM-DD)", "Appointment Date", "Billing Date (YYYY-MM-DD)"]:
        if col in template_df.columns:
            template_df[col] = pd.to_datetime(
                template_df[col],
                errors="coerce"
            ).dt.strftime("%Y-%m-%d").fillna("")

    return template_df


def normalize_yes_no(value):
    text_value = clean_text(value).lower()

    if text_value in ["yes", "y", "true", "1", "done", "billed"]:
        return "Yes"

    if text_value in ["no", "n", "false", "0", "not done", "pending", ""]:
        return "No"

    return ""


def parse_upload_date(value):
    value = clean_text(value)

    if value == "":
        return "", ""

    parsed_date = pd.to_datetime(
        value,
        errors="coerce",
        dayfirst=False
    )

    if pd.isna(parsed_date):
        return "", f"Invalid date '{value}'. Use YYYY-MM-DD."

    return parsed_date.strftime("%Y-%m-%d"), ""


def apply_billing_update_upload(uploaded_df):
    update_df = normalize_columns(uploaded_df)

    column_aliases = {
        "Tracker ID": ["Tracker ID", "ID", "id"],
        "FCN": ["FCN", "fcn"],
        "Appointment Date": ["Appointment Date", "Appointment Date (YYYY-MM-DD)", "appointment_date"],
        "Invoice No.": ["Invoice No.", "Invoice No", "Invoice Number", "invoice_no"],
        "Billing Date (YYYY-MM-DD)": ["Billing Date (YYYY-MM-DD)", "Billing Date", "billing_date"],
        "Billed Qty.": ["Billed Qty.", "Billed Qty", "Billing Qty", "billed_qty"],
        "Sent For Billing (Yes/No)": ["Sent For Billing (Yes/No)", "Sent For Billing", "sent_for_billing"],
        "Billing Done (Yes/No)": ["Billing Done (Yes/No)", "Billing Done", "billing_done"],
        "Remark": ["Remark", "Remarks", "remark"],
    }

    rename_map = {}

    for target_col, aliases in column_aliases.items():
        found_col = first_existing_column(update_df, aliases)
        if found_col:
            rename_map[found_col] = target_col

    update_df = update_df.rename(columns=rename_map)

    required_cols = ["Tracker ID"]
    missing_cols = [col for col in required_cols if col not in update_df.columns]

    if missing_cols:
        return {
            "updated_rows": 0,
            "error_rows": pd.DataFrame([{
                "Error": f"Missing required columns: {missing_cols}"
            }])
        }

    for col in column_aliases:
        if col not in update_df.columns:
            update_df[col] = ""

    tracker_df = get_tracker_df()

    if tracker_df.empty:
        return {
            "updated_rows": 0,
            "error_rows": pd.DataFrame([{"Error": "No tracker rows found."}])
        }

    tracker_by_id = {
        int(row["id"]): row
        for _, row in tracker_df.iterrows()
    }

    updated_rows = 0
    updated_ids = []
    error_rows = []

    for row_no, row in update_df.iterrows():
        try:
            tracker_id = int(clean_number(row["Tracker ID"]))
        except Exception:
            error_rows.append({
                "Excel Row": row_no + 2,
                "Tracker ID": row.get("Tracker ID", ""),
                "Error": "Invalid Tracker ID"
            })
            continue

        if tracker_id not in tracker_by_id:
            error_rows.append({
                "Excel Row": row_no + 2,
                "Tracker ID": tracker_id,
                "Error": "Tracker ID not found"
            })
            continue

        tracker_row = tracker_by_id[tracker_id]
        allocated_qty = clean_number(tracker_row.get("allocated_qty", 0))
        fcn = clean_text(row["FCN"])
        appointment_date, appointment_date_error = parse_upload_date(row["Appointment Date"])
        billed_qty = clean_number(row["Billed Qty."])
        sent_for_billing = normalize_yes_no(row["Sent For Billing (Yes/No)"])
        billing_done = normalize_yes_no(row["Billing Done (Yes/No)"])
        invoice_no = clean_text(row["Invoice No."])
        remark = clean_text(row["Remark"])
        billing_date, date_error = parse_upload_date(row["Billing Date (YYYY-MM-DD)"])

        if appointment_date_error:
            error_rows.append({
                "Excel Row": row_no + 2,
                "Tracker ID": tracker_id,
                "Error": appointment_date_error
            })
            continue

        if date_error:
            error_rows.append({
                "Excel Row": row_no + 2,
                "Tracker ID": tracker_id,
                "Error": date_error
            })
            continue

        if billing_done == "":
            error_rows.append({
                "Excel Row": row_no + 2,
                "Tracker ID": tracker_id,
                "Error": "Billing Done must be Yes or No"
            })
            continue

        if sent_for_billing == "":
            error_rows.append({
                "Excel Row": row_no + 2,
                "Tracker ID": tracker_id,
                "Error": "Sent For Billing must be Yes or No"
            })
            continue

        if billing_done == "Yes" and invoice_no == "":
            error_rows.append({
                "Excel Row": row_no + 2,
                "Tracker ID": tracker_id,
                "Error": "Invoice No. is required when Billing Done is Yes"
            })
            continue

        if billing_done == "Yes" and billing_date == "":
            error_rows.append({
                "Excel Row": row_no + 2,
                "Tracker ID": tracker_id,
                "Error": "Billing Date is required when Billing Done is Yes"
            })
            continue

        if billing_done == "Yes" and billed_qty <= 0:
            billed_qty = allocated_qty

        if billing_done == "Yes" or billed_qty > 0 or invoice_no != "":
            sent_for_billing = "Yes"

        billed_qty = min(billed_qty, allocated_qty)
        balance_to_bill = max(allocated_qty - billed_qty, 0)

        if billed_qty > 0 and billing_done == "No" and balance_to_bill <= 0:
            billing_done = "Yes"

        update_fields = {
            "sent_for_billing": sent_for_billing,
            "fcn": fcn,
            "appointment_date": appointment_date,
            "invoice_no": invoice_no,
            "billing_date": billing_date,
            "billing_done": billing_done,
            "remark": remark,
            "id": tracker_id,
        }

        if "billed_qty" in tracker_df.columns and "balance_to_bill" in tracker_df.columns:
            db_execute("""
            UPDATE allocation_tracker
            SET
                sent_for_billing = :sent_for_billing,
                fcn = :fcn,
                appointment_date = :appointment_date,
                invoice_no = :invoice_no,
                billing_date = :billing_date,
                billing_done = :billing_done,
                billed_qty = :billed_qty,
                balance_to_bill = :balance_to_bill,
                remark = :remark
            WHERE id = :id
            """, {
                **update_fields,
                "billed_qty": billed_qty,
                "balance_to_bill": balance_to_bill,
            }, clear_cache=False)

        else:
            db_execute("""
            UPDATE allocation_tracker
            SET
                sent_for_billing = :sent_for_billing,
                fcn = :fcn,
                appointment_date = :appointment_date,
                invoice_no = :invoice_no,
                billing_date = :billing_date,
                billing_done = :billing_done,
                remark = :remark
            WHERE id = :id
            """, update_fields, clear_cache=False)

        updated_rows += 1
        updated_ids.append(tracker_id)

    if updated_rows > 0:
        st.cache_data.clear()

    return {
        "updated_rows": updated_rows,
        "updated_ids": updated_ids,
        "error_rows": pd.DataFrame(error_rows)
    }


def verify_sales_billing_against_tracker(uploaded_df, tracker_df):
    sales_df = normalize_columns(uploaded_df)

    aliases = {
        "PO No.": ["PO No", "PO No.", "PO", "PONo"],
        "FSN": ["FSN"],
        "SAP Code": ["Item", "SAP Code", "Item Code"],
        "RR Warehouse": ["Site", "RR Warehouse", "RR WH"],
        "Invoice No.": ["GST Invoice No", "GST Invoice No.", "Invoice No.", "Invoice No"],
        "Invoice Date": ["Invoice Date", "Billing Date"],
        "Dispatch Qty": ["Dispatch Qty", "Dispatch Qty.", "Billed Qty.", "Billing Qty"],
        "Dispatch Amount": ["Dispatch Amount", "Dispatch Amount."],
        "Sales Id": ["Sales Id", "Sales ID", "Order ID"],
        "Item Name": ["Item Name", "Title"],
    }

    rename_map = {}

    for target_col, options in aliases.items():
        found_col = first_existing_column(sales_df, options)
        if found_col:
            rename_map[found_col] = target_col

    sales_df = sales_df.rename(columns=rename_map)

    required_cols = [
        "PO No.",
        "FSN",
        "SAP Code",
        "RR Warehouse",
        "Invoice No.",
        "Invoice Date",
        "Dispatch Qty",
    ]
    missing_cols = [col for col in required_cols if col not in sales_df.columns]

    if missing_cols:
        return {
            "summary": pd.DataFrame([{"Error": f"Missing required sales columns: {missing_cols}"}]),
            "matches": pd.DataFrame(),
            "exceptions": pd.DataFrame(),
            "unmatched_sales": pd.DataFrame(),
        }

    for col in ["PO No.", "FSN", "SAP Code", "RR Warehouse", "Invoice No."]:
        sales_df[col] = sales_df[col].apply(clean_text)

    for col in ["PO No.", "FSN", "SAP Code", "RR Warehouse"]:
        sales_df[col] = sales_df[col].str.upper()

    sales_df["Dispatch Qty"] = sales_df["Dispatch Qty"].apply(clean_number)
    sales_df["Invoice Date"] = pd.to_datetime(
        sales_df["Invoice Date"],
        errors="coerce"
    ).dt.strftime("%Y-%m-%d").fillna("")

    if "Dispatch Amount" in sales_df.columns:
        sales_df["Dispatch Amount"] = sales_df["Dispatch Amount"].apply(clean_number)
    else:
        sales_df["Dispatch Amount"] = 0

    positive_sales_df = sales_df[sales_df["Dispatch Qty"] > 0].copy()

    key_cols = ["PO No.", "FSN", "SAP Code", "RR Warehouse"]

    sales_grouped = positive_sales_df.groupby(
        key_cols,
        dropna=False
    ).agg(
        sales_billed_qty=("Dispatch Qty", "sum"),
        sales_dispatch_amount=("Dispatch Amount", "sum"),
        sales_invoice_no=("Invoice No.", lambda values: ", ".join(
            sorted(set(clean_text(v) for v in values if clean_text(v) != ""))
        )),
        sales_invoice_count=("Invoice No.", "nunique"),
        latest_sales_invoice_date=("Invoice Date", "max"),
        sales_rows=("FSN", "size"),
    ).reset_index()

    tracker_norm = tracker_df.copy()

    for col in ["po_no", "fsn", "sap_code", "rr_warehouse"]:
        if col not in tracker_norm.columns:
            tracker_norm[col] = ""
        tracker_norm[col] = tracker_norm[col].apply(clean_text).str.upper()

    for col in ["allocated_qty", "billed_qty", "balance_to_bill"]:
        if col not in tracker_norm.columns:
            tracker_norm[col] = 0
        tracker_norm[col] = tracker_norm[col].apply(clean_number)

    tracker_norm["tracker_key_count"] = tracker_norm.groupby(
        ["po_no", "fsn", "sap_code", "rr_warehouse"]
    )["id"].transform("count")

    compare_df = tracker_norm.merge(
        sales_grouped,
        how="left",
        left_on=["po_no", "fsn", "sap_code", "rr_warehouse"],
        right_on=["PO No.", "FSN", "SAP Code", "RR Warehouse"]
    )

    compare_df["sales_billed_qty"] = compare_df["sales_billed_qty"].fillna(0)
    compare_df["sales_dispatch_amount"] = compare_df["sales_dispatch_amount"].fillna(0)
    compare_df["sales_invoice_no"] = compare_df["sales_invoice_no"].fillna("")
    compare_df["latest_sales_invoice_date"] = compare_df["latest_sales_invoice_date"].fillna("")
    compare_df["sales_invoice_count"] = compare_df["sales_invoice_count"].fillna(0).astype(int)
    compare_df["sales_rows"] = compare_df["sales_rows"].fillna(0).astype(int)
    compare_df["qty_difference"] = compare_df["billed_qty"] - compare_df["sales_billed_qty"]
    compare_df["sales_balance_to_bill"] = (
        compare_df["allocated_qty"] - compare_df["sales_billed_qty"]
    ).apply(lambda value: max(value, 0))

    def verification_status(row):
        if clean_number(row["sales_billed_qty"]) <= 0 and clean_number(row["billed_qty"]) <= 0:
            return "No Billing In Tracker Or Sales"
        if clean_number(row["sales_billed_qty"]) <= 0 and clean_number(row["billed_qty"]) > 0:
            return "Tracker Billed But Sales Not Found"
        if clean_number(row["sales_billed_qty"]) > clean_number(row["allocated_qty"]):
            return "Sales Qty More Than Allocated"
        if clean_number(row["tracker_key_count"]) > 1:
            return "Duplicate Tracker Key - Review Before Auto Update"
        if clean_number(row["qty_difference"]) == 0:
            return "Matched"
        return "Ready To Update From Sales"

    compare_df["Verification Status"] = compare_df.apply(verification_status, axis=1)

    match_columns = [
        "id",
        "po_no",
        "fsn",
        "sap_code",
        "rr_warehouse",
        "allocated_qty",
        "billed_qty",
        "sales_billed_qty",
        "sales_balance_to_bill",
        "qty_difference",
        "sales_invoice_no",
        "latest_sales_invoice_date",
        "sales_invoice_count",
        "sales_rows",
        "tracker_key_count",
        "billing_done",
        "Verification Status",
    ]
    match_columns = [col for col in match_columns if col in compare_df.columns]

    comparison_output = compare_df[match_columns].rename(columns={
        "id": "Tracker ID",
        "po_no": "PO No.",
        "fsn": "FSN",
        "sap_code": "SAP Code",
        "rr_warehouse": "RR Warehouse",
        "allocated_qty": "Allocated Qty.",
        "billed_qty": "Tracker Billed Qty.",
        "sales_billed_qty": "Sales Billed Qty.",
        "sales_balance_to_bill": "Sales Balance To Bill",
        "qty_difference": "Tracker - Sales Diff",
        "sales_invoice_no": "Sales Invoice No.",
        "latest_sales_invoice_date": "Latest Sales Invoice Date",
        "sales_invoice_count": "Invoice Count",
        "sales_rows": "Sales Rows",
        "tracker_key_count": "Tracker Key Count",
        "billing_done": "Tracker Billing Done",
    })

    matches = comparison_output[
        comparison_output["Verification Status"] == "Matched"
    ].copy()

    exceptions = comparison_output[
        comparison_output["Verification Status"].isin([
            "Tracker Billed But Sales Not Found",
            "Sales Qty More Than Allocated",
            "Duplicate Tracker Key - Review Before Auto Update",
        ])
    ].copy()

    updateable_rows = comparison_output[
        comparison_output["Verification Status"].isin([
            "Matched",
            "Ready To Update From Sales",
        ])
    ].copy()

    sales_matched = sales_grouped.merge(
        tracker_norm[["id", "po_no", "fsn", "sap_code", "rr_warehouse"]],
        how="left",
        left_on=["PO No.", "FSN", "SAP Code", "RR Warehouse"],
        right_on=["po_no", "fsn", "sap_code", "rr_warehouse"]
    )

    unmatched_sales = sales_matched[sales_matched["id"].isna()].copy()
    unmatched_sales = unmatched_sales[
        [
            "PO No.",
            "FSN",
            "SAP Code",
            "RR Warehouse",
            "sales_billed_qty",
            "sales_dispatch_amount",
            "sales_invoice_no",
            "latest_sales_invoice_date",
            "sales_invoice_count",
            "sales_rows",
        ]
    ].rename(columns={
        "sales_billed_qty": "Sales Billed Qty.",
        "sales_dispatch_amount": "Sales Dispatch Amount",
        "sales_invoice_no": "Sales Invoice No.",
        "latest_sales_invoice_date": "Latest Sales Invoice Date",
        "sales_invoice_count": "Invoice Count",
        "sales_rows": "Sales Rows",
    })

    summary = pd.DataFrame([
        {"Metric": "Sales file total rows", "Value": len(sales_df)},
        {"Metric": "Positive dispatch rows used", "Value": len(positive_sales_df)},
        {"Metric": "Negative/zero dispatch rows ignored", "Value": len(sales_df) - len(positive_sales_df)},
        {"Metric": "Grouped sales billing keys", "Value": len(sales_grouped)},
        {"Metric": "Tracker rows checked", "Value": len(tracker_norm)},
        {"Metric": "Matched rows", "Value": len(matches)},
        {"Metric": "Rows safe for sales update", "Value": len(updateable_rows)},
        {"Metric": "Exception rows", "Value": len(exceptions)},
        {"Metric": "Unmatched sales billing keys", "Value": len(unmatched_sales)},
        {"Metric": "Tracker billed qty total", "Value": tracker_norm["billed_qty"].sum()},
        {"Metric": "Matched sales billed qty total", "Value": comparison_output["Sales Billed Qty."].sum() if not comparison_output.empty else 0},
    ])

    return {
        "summary": summary,
        "matches": matches,
        "updateable_rows": updateable_rows,
        "exceptions": exceptions,
        "unmatched_sales": unmatched_sales,
    }


def apply_sales_billing_update_upload(uploaded_df, tracker_df):
    verification_result = verify_sales_billing_against_tracker(
        uploaded_df,
        tracker_df
    )
    updateable_rows = verification_result["updateable_rows"]

    if updateable_rows.empty:
        return {
            **verification_result,
            "updated_rows": 0,
            "updated_ids": [],
        }

    tracker_columns = get_allocation_tracker_columns()
    updated_rows = 0
    updated_ids = []

    for _, row in updateable_rows.iterrows():
        tracker_id = int(clean_number(row["Tracker ID"]))
        allocated_qty = clean_number(row["Allocated Qty."])
        billed_qty = clean_number(row["Sales Billed Qty."])
        balance_to_bill = max(allocated_qty - billed_qty, 0)
        billing_done = "Yes" if balance_to_bill <= 0 and allocated_qty > 0 else "No"

        params = {
            "id": tracker_id,
            "sent_for_billing": "Yes",
            "invoice_no": clean_text(row["Sales Invoice No."]),
            "billing_date": clean_text(row["Latest Sales Invoice Date"]),
            "billing_done": billing_done,
            "billed_qty": billed_qty,
            "balance_to_bill": balance_to_bill,
            "billing_source": "Sales Billing Upload",
        }

        update_fields = [
            "sent_for_billing = :sent_for_billing",
            "invoice_no = :invoice_no",
            "billing_date = :billing_date",
            "billing_done = :billing_done",
        ]

        if "billed_qty" in tracker_columns:
            update_fields.append("billed_qty = :billed_qty")

        if "balance_to_bill" in tracker_columns:
            update_fields.append("balance_to_bill = :balance_to_bill")

        if "billing_source" in tracker_columns:
            update_fields.append("billing_source = :billing_source")

        db_execute(f"""
        UPDATE allocation_tracker
        SET {", ".join(update_fields)}
        WHERE id = :id
        """, params, clear_cache=False)

        updated_rows += 1
        updated_ids.append(tracker_id)

    if updated_rows > 0:
        st.cache_data.clear()

    return {
        **verification_result,
        "updated_rows": updated_rows,
        "updated_ids": updated_ids,
    }


def prepare_appointment_upload(uploaded_df):
    upload_df = normalize_columns(uploaded_df)

    aliases = {
        "Appointment No.": ["Appointment No.", "Appointment No", "Appointment Number"],
        "Appointment Date": ["Appointment Date", "Delivery Date"],
        "PO No.": ["PO No.", "PO No", "PONo", "PO"],
        "FSN": ["FSN"],
        "FK Warehouse": ["FK Warehouse", "FK FC"],
        "Appointment Qty.": ["Appointment Qty.", "Appointment Qty", "Qty", "Quantity"],
        "Remark": ["Remark", "Remarks"],
    }

    rename_map = {}

    for target_col, options in aliases.items():
        found_col = first_existing_column(upload_df, options)
        if found_col:
            rename_map[found_col] = target_col

    upload_df = upload_df.rename(columns=rename_map)
    required_cols = ["Appointment No.", "Appointment Date", "FSN", "FK Warehouse", "Appointment Qty."]
    missing_cols = [col for col in required_cols if col not in upload_df.columns]

    if missing_cols:
        return pd.DataFrame(), missing_cols

    for col in ["PO No.", "Remark"]:
        if col not in upload_df.columns:
            upload_df[col] = ""

    for col in ["Appointment No.", "PO No.", "FSN", "FK Warehouse", "Remark"]:
        upload_df[col] = upload_df[col].apply(clean_text)

    upload_df["FK Warehouse"] = upload_df["FK Warehouse"].str.upper()
    upload_df["Appointment Qty."] = upload_df["Appointment Qty."].apply(clean_number)
    upload_df["Appointment Date"] = upload_df["Appointment Date"].apply(
        lambda value: parse_upload_date(value)[0]
    )

    upload_df = upload_df[
        (upload_df["Appointment No."] != "") &
        (upload_df["Appointment Date"] != "") &
        (upload_df["FSN"] != "") &
        (upload_df["FK Warehouse"] != "") &
        (upload_df["Appointment Qty."] > 0)
    ]

    return upload_df[
        [
            "Appointment No.",
            "Appointment Date",
            "PO No.",
            "FSN",
            "FK Warehouse",
            "Appointment Qty.",
            "Remark",
        ]
    ], []


def save_appointment_upload(appointment_df):
    saved_count = 0

    for _, row in appointment_df.iterrows():
        params = {
            "appointment_no": row["Appointment No."],
            "appointment_date": row["Appointment Date"],
            "po_no": row["PO No."],
            "fsn": row["FSN"],
            "fk_warehouse": row["FK Warehouse"],
            "appointment_qty": clean_number(row["Appointment Qty."]),
            "remark": row["Remark"],
            "upload_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

        db_execute("""
        DELETE FROM appointment_tracker
        WHERE appointment_no = :appointment_no
        AND fsn = :fsn
        AND COALESCE(fk_warehouse, '') = COALESCE(:fk_warehouse, '')
        AND COALESCE(po_no, '') = COALESCE(:po_no, '')
        """, params, clear_cache=False)

        db_execute("""
        INSERT INTO appointment_tracker (
            upload_date,
            appointment_no,
            appointment_date,
            po_no,
            fsn,
            fk_warehouse,
            appointment_qty,
            remark
        )
        VALUES (
            :upload_date,
            :appointment_no,
            :appointment_date,
            :po_no,
            :fsn,
            :fk_warehouse,
            :appointment_qty,
            :remark
        )
        """, params, clear_cache=False)

        saved_count += 1

    if saved_count > 0:
        st.cache_data.clear()

    return saved_count


def get_manual_allocation_template_df():
    return pd.DataFrame({
        "PO No.": ["PO123"],
        "Order ID": ["ORDER123"],
        "Buyer Code": ["BUYER01"],
        "FSN": ["F05JODEAVB48E1"],
        "Title": ["Product title"],
        "RR Warehouse": ["TSHYD01"],
        "FK Warehouse": ["hyderabad_medchal_01"],
        "SAP Code": ["SAP001"],
        "Pending Amount": [0],
        "Allocated Qty.": [10],
        "Appointment No.": [""],
        "Appointment Date": [""],
        "Remark": ["Manual emergency allocation"],
    })


def prepare_manual_allocation_upload(uploaded_df):
    manual_df = normalize_columns(uploaded_df)

    aliases = {
        "PO No.": ["PO No.", "PO No", "PONo"],
        "Order ID": ["Order ID", "Order Id"],
        "Buyer Code": ["Buyer Code"],
        "FSN": ["FSN"],
        "Title": ["Title", "Item Description"],
        "RR Warehouse": ["RR Warehouse", "Site", "Site Id"],
        "FK Warehouse": ["FK Warehouse", "FK FC"],
        "SAP Code": ["SAP Code", "Item Code", "Item"],
        "Pending Amount": ["Pending Amount", "Pending Amt"],
        "Allocated Qty.": ["Allocated Qty.", "Allocated Qty", "Qty"],
        "Appointment No.": ["Appointment No.", "Appointment No"],
        "Appointment Date": ["Appointment Date"],
        "Remark": ["Remark", "Remarks"],
    }

    rename_map = {}

    for target_col, options in aliases.items():
        found_col = first_existing_column(manual_df, options)
        if found_col:
            rename_map[found_col] = target_col

    manual_df = manual_df.rename(columns=rename_map)
    required_cols = ["PO No.", "FSN", "Title", "RR Warehouse", "FK Warehouse", "SAP Code", "Allocated Qty."]
    missing_cols = [col for col in required_cols if col not in manual_df.columns]

    if missing_cols:
        return pd.DataFrame(), missing_cols

    for col in ["Order ID", "Buyer Code", "Pending Amount", "Appointment No.", "Appointment Date", "Remark"]:
        if col not in manual_df.columns:
            manual_df[col] = "" if col != "Pending Amount" else 0

    for col in ["PO No.", "Order ID", "Buyer Code", "FSN", "Title", "RR Warehouse", "FK Warehouse", "SAP Code", "Appointment No.", "Appointment Date", "Remark"]:
        manual_df[col] = manual_df[col].apply(clean_text)

    manual_df["RR Warehouse"] = manual_df["RR Warehouse"].str.upper()
    manual_df["Allocated Qty."] = manual_df["Allocated Qty."].apply(clean_number)
    manual_df["Pending Amount"] = manual_df["Pending Amount"].apply(clean_number)

    manual_df = manual_df[
        (manual_df["FSN"] != "") &
        (manual_df["Allocated Qty."] > 0)
    ]

    return manual_df[
        [
            "PO No.",
            "Order ID",
            "Buyer Code",
            "FSN",
            "Title",
            "RR Warehouse",
            "FK Warehouse",
            "SAP Code",
            "Pending Amount",
            "Allocated Qty.",
            "Appointment No.",
            "Appointment Date",
            "Remark",
        ]
    ], []


def save_manual_allocation_upload(manual_df):
    saved_count = 0
    tracker_columns = get_allocation_tracker_columns()

    for _, row in manual_df.iterrows():
        params = {
            "allocation_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "po_no": row["PO No."],
            "order_id": row["Order ID"],
            "buyer_code": row["Buyer Code"],
            "fsn": row["FSN"],
            "title": row["Title"],
            "rr_warehouse": row["RR Warehouse"],
            "fk_warehouse": row["FK Warehouse"],
            "sap_code": row["SAP Code"],
            "pending_amount": clean_number(row["Pending Amount"]),
            "allocated_qty": clean_number(row["Allocated Qty."]),
            "appointment_no": row["Appointment No."],
            "appointment_date": row["Appointment Date"],
            "sent_for_billing": "No",
            "billing_done": "No",
            "remark": row["Remark"] or "Manual Emergency Allocation",
            "billed_qty": 0,
            "balance_to_bill": clean_number(row["Allocated Qty."]),
            "billing_source": "Manual Allocation Upload",
        }

        insert_columns = [
            "allocation_date", "po_no", "fsn", "title", "rr_warehouse",
            "fk_warehouse", "sap_code", "allocated_qty", "sent_for_billing",
            "billing_done", "remark"
        ]

        optional_columns = [
            "order_id", "buyer_code", "pending_amount", "appointment_no",
            "appointment_date", "billed_qty", "balance_to_bill", "billing_source"
        ]

        for col in optional_columns:
            if col in tracker_columns:
                insert_columns.append(col)

        value_keys = [f":{col}" for col in insert_columns]

        db_execute(f"""
        INSERT INTO allocation_tracker ({", ".join(insert_columns)})
        VALUES ({", ".join(value_keys)})
        """, params, clear_cache=False)

        saved_count += 1

    if saved_count > 0:
        st.cache_data.clear()

    return saved_count


def apply_delete_allocation_upload(uploaded_df):
    delete_df = normalize_columns(uploaded_df)

    tracker_col = first_existing_column(delete_df, ["Tracker ID", "ID", "id"])
    reason_col = first_existing_column(delete_df, ["Delete Reason", "Reason", "Remark"])

    if tracker_col is None:
        return 0, pd.DataFrame([{"Error": "Missing Tracker ID column"}])

    deleted_count = 0
    errors = []

    for row_no, row in delete_df.iterrows():
        tracker_id = int(clean_number(row[tracker_col]))
        reason = clean_text(row[reason_col]) if reason_col else ""

        if tracker_id <= 0:
            errors.append({
                "Excel Row": row_no + 2,
                "Tracker ID": row.get(tracker_col, ""),
                "Error": "Invalid Tracker ID"
            })
            continue

        existing = db_read(
            "SELECT id FROM allocation_tracker WHERE id = :id",
            {"id": tracker_id},
            use_cache=False
        )

        if existing.empty:
            errors.append({
                "Excel Row": row_no + 2,
                "Tracker ID": tracker_id,
                "Error": "Tracker ID not found"
            })
            continue

        db_execute(
            "DELETE FROM allocation_tracker WHERE id = :id",
            {"id": tracker_id},
            clear_cache=False
        )
        log_activity("delete_allocation", f"Deleted allocation ID {tracker_id}. Reason: {reason}")
        deleted_count += 1

    if deleted_count > 0:
        st.cache_data.clear()

    return deleted_count, pd.DataFrame(errors)


def get_tracker_correction_template_df(tracker_df):
    template_df = tracker_df.copy()

    export_columns = [
        "id",
        "allocation_date",
        "po_no",
        "order_id",
        "buyer_code",
        "fsn",
        "title",
        "rr_warehouse",
        "fk_warehouse",
        "sap_code",
        "fcn",
        "appointment_no",
        "appointment_date",
        "pending_amount",
        "allocated_qty",
        "billed_qty",
        "balance_to_bill",
        "sent_for_billing",
        "sent_date",
        "invoice_no",
        "billing_date",
        "billing_done",
        "remark",
    ]

    export_columns = [col for col in export_columns if col in template_df.columns]
    template_df = template_df[export_columns]

    template_df.insert(0, "Action", "Replace")

    template_df = template_df.rename(columns={
        "id": "Tracker ID",
        "allocation_date": "Allocation Date",
        "po_no": "PO No.",
        "order_id": "Order ID",
        "buyer_code": "Buyer Code",
        "fsn": "FSN",
        "title": "Title",
        "rr_warehouse": "RR Warehouse",
        "fk_warehouse": "FK Warehouse",
        "sap_code": "SAP Code",
        "fcn": "FCN",
        "appointment_no": "Appointment No.",
        "appointment_date": "Appointment Date",
        "pending_amount": "Pending Amount",
        "allocated_qty": "Allocated Qty.",
        "billed_qty": "Billed Qty.",
        "balance_to_bill": "Remaining Qty.",
        "sent_for_billing": "Sent For Billing (Yes/No)",
        "sent_date": "Sent Date (YYYY-MM-DD)",
        "invoice_no": "Invoice No.",
        "billing_date": "Billing Date (YYYY-MM-DD)",
        "billing_done": "Billing Done (Yes/No)",
        "remark": "Remark",
    })

    for col in ["Allocation Date", "Appointment Date", "Sent Date (YYYY-MM-DD)", "Billing Date (YYYY-MM-DD)"]:
        if col in template_df.columns:
            template_df[col] = pd.to_datetime(
                template_df[col],
                errors="coerce"
            ).dt.strftime("%Y-%m-%d").fillna("")

    new_row = {col: "" for col in template_df.columns}
    new_row["Action"] = "New"

    if "Allocated Qty." in new_row:
        new_row["Allocated Qty."] = 0

    if "Billed Qty." in new_row:
        new_row["Billed Qty."] = 0

    if "Remaining Qty." in new_row:
        new_row["Remaining Qty."] = 0

    template_df = pd.concat(
        [pd.DataFrame([new_row]), template_df],
        ignore_index=True
    )

    return template_df


def apply_tracker_correction_upload(uploaded_df):
    correction_df = normalize_columns(uploaded_df)

    aliases = {
        "Action": ["Action", "Correction Action", "Mode"],
        "Tracker ID": ["Tracker ID", "ID", "id"],
        "Allocation Date": ["Allocation Date", "allocation_date"],
        "PO No.": ["PO No.", "PO No", "PONo", "po_no"],
        "Order ID": ["Order ID", "Order Id", "order_id"],
        "Buyer Code": ["Buyer Code", "buyer_code"],
        "FSN": ["FSN", "fsn"],
        "Title": ["Title", "Item Description", "title"],
        "RR Warehouse": ["RR Warehouse", "RR WH", "Site", "rr_warehouse"],
        "FK Warehouse": ["FK Warehouse", "FK FC", "fk_warehouse"],
        "SAP Code": ["SAP Code", "Item Code", "sap_code"],
        "FCN": ["FCN", "fcn"],
        "Appointment No.": ["Appointment No.", "Appointment No", "appointment_no"],
        "Appointment Date": ["Appointment Date", "appointment_date"],
        "Pending Amount": ["Pending Amount", "Pending Amt", "pending_amount"],
        "Allocated Qty.": ["Allocated Qty.", "Allocated Qty", "allocated_qty"],
        "Billed Qty.": ["Billed Qty.", "Billed Qty", "Billing Qty", "billed_qty"],
        "Remaining Qty.": ["Remaining Qty.", "Remaining Qty", "Balance To Bill", "balance_to_bill"],
        "Sent For Billing (Yes/No)": ["Sent For Billing (Yes/No)", "Sent For Billing", "sent_for_billing"],
        "Sent Date (YYYY-MM-DD)": ["Sent Date (YYYY-MM-DD)", "Sent Date", "sent_date"],
        "Invoice No.": ["Invoice No.", "Invoice No", "Invoice Number", "invoice_no"],
        "Billing Date (YYYY-MM-DD)": ["Billing Date (YYYY-MM-DD)", "Billing Date", "billing_date"],
        "Billing Done (Yes/No)": ["Billing Done (Yes/No)", "Billing Done", "billing_done"],
        "Remark": ["Remark", "Remarks", "remark"],
    }

    rename_map = {}

    for target_col, options in aliases.items():
        found_col = first_existing_column(correction_df, options)
        if found_col:
            rename_map[found_col] = target_col

    correction_df = correction_df.rename(columns=rename_map)

    if "Tracker ID" not in correction_df.columns:
        correction_df["Tracker ID"] = ""

    if "Action" not in correction_df.columns:
        correction_df["Action"] = "Replace"

    tracker_columns = get_allocation_tracker_columns()

    tracker_df = get_tracker_df()
    tracker_by_id = {
        int(row["id"]): row
        for _, row in tracker_df.iterrows()
    }

    replace_map = {
        "allocation_date": "Allocation Date",
        "po_no": "PO No.",
        "order_id": "Order ID",
        "buyer_code": "Buyer Code",
        "fsn": "FSN",
        "title": "Title",
        "rr_warehouse": "RR Warehouse",
        "fk_warehouse": "FK Warehouse",
        "sap_code": "SAP Code",
        "fcn": "FCN",
        "appointment_no": "Appointment No.",
        "appointment_date": "Appointment Date",
        "pending_amount": "Pending Amount",
        "allocated_qty": "Allocated Qty.",
        "billed_qty": "Billed Qty.",
        "balance_to_bill": "Remaining Qty.",
        "sent_for_billing": "Sent For Billing (Yes/No)",
        "sent_date": "Sent Date (YYYY-MM-DD)",
        "invoice_no": "Invoice No.",
        "billing_date": "Billing Date (YYYY-MM-DD)",
        "billing_done": "Billing Done (Yes/No)",
        "remark": "Remark",
    }

    numeric_columns = {"pending_amount", "allocated_qty", "billed_qty", "balance_to_bill"}
    date_columns = {"allocation_date", "appointment_date", "sent_date", "billing_date"}
    yes_no_columns = {"sent_for_billing", "billing_done"}

    replaced_rows = 0
    deleted_rows = 0
    inserted_rows = 0
    errors = []

    for row_no, row in correction_df.iterrows():
        action = clean_text(row.get("Action", "Replace")).lower()

        if action in ["delete", "remove", "deleted"]:
            action = "delete"
        elif action in ["new", "insert", "add", "create"]:
            action = "new"
        elif action in ["replace", "update", "correct", "correction", ""]:
            action = "replace"
        else:
            errors.append({
                "Excel Row": row_no + 2,
                "Tracker ID": row.get("Tracker ID", ""),
                "Error": "Action must be New, Replace, or Delete"
            })
            continue

        tracker_id = int(clean_number(row["Tracker ID"]))

        if action != "new" and tracker_id <= 0:
            errors.append({
                "Excel Row": row_no + 2,
                "Tracker ID": row.get("Tracker ID", ""),
                "Error": "Tracker ID is required for Replace/Delete"
            })
            continue

        if action != "new" and tracker_id not in tracker_by_id:
            errors.append({
                "Excel Row": row_no + 2,
                "Tracker ID": tracker_id,
                "Error": "Tracker ID not found"
            })
            continue

        if action == "new":
            required_new_cols = [
                "PO No.",
                "FSN",
                "Title",
                "RR Warehouse",
                "FK Warehouse",
                "SAP Code",
                "Allocated Qty.",
            ]
            missing_new_cols = [
                col
                for col in required_new_cols
                if col not in correction_df.columns or clean_text(row.get(col, "")) == ""
            ]

            if missing_new_cols:
                errors.append({
                    "Excel Row": row_no + 2,
                    "Tracker ID": row.get("Tracker ID", ""),
                    "Error": f"New row missing required columns: {missing_new_cols}"
                })
                continue

            new_values = {}
            insert_columns = []

            for db_col, upload_col in replace_map.items():
                if db_col not in tracker_columns or upload_col not in correction_df.columns:
                    continue

                raw_value = row.get(upload_col, "")

                if db_col in numeric_columns:
                    value = clean_number(raw_value)
                elif db_col in date_columns:
                    value, date_error = parse_upload_date(raw_value)

                    if date_error:
                        errors.append({
                            "Excel Row": row_no + 2,
                            "Tracker ID": row.get("Tracker ID", ""),
                            "Error": date_error
                        })
                        break
                elif db_col in yes_no_columns:
                    value = normalize_yes_no(raw_value)

                    if value == "":
                        value = "No"
                else:
                    value = clean_text(raw_value)

                new_values[db_col] = value
                insert_columns.append(db_col)

            else:
                allocated_qty = clean_number(new_values.get("allocated_qty", 0))
                billed_qty = clean_number(new_values.get("billed_qty", 0))

                if allocated_qty <= 0:
                    errors.append({
                        "Excel Row": row_no + 2,
                        "Tracker ID": row.get("Tracker ID", ""),
                        "Error": "Allocated Qty. must be greater than 0 for New rows"
                    })
                    continue

                if billed_qty < 0 or billed_qty > allocated_qty:
                    errors.append({
                        "Excel Row": row_no + 2,
                        "Tracker ID": row.get("Tracker ID", ""),
                        "Error": "Billed Qty. cannot be negative or greater than Allocated Qty."
                    })
                    continue

                if "allocation_date" in tracker_columns and clean_text(new_values.get("allocation_date", "")) == "":
                    new_values["allocation_date"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                    if "allocation_date" not in insert_columns:
                        insert_columns.append("allocation_date")

                if "balance_to_bill" in tracker_columns:
                    remaining_raw = clean_text(row.get("Remaining Qty.", ""))
                    new_values["balance_to_bill"] = (
                        clean_number(row.get("Remaining Qty.", 0))
                        if remaining_raw != ""
                        else max(allocated_qty - billed_qty, 0)
                    )

                    if "balance_to_bill" not in insert_columns:
                        insert_columns.append("balance_to_bill")

                balance_to_bill = clean_number(new_values.get("balance_to_bill", max(allocated_qty - billed_qty, 0)))

                if balance_to_bill < 0:
                    errors.append({
                        "Excel Row": row_no + 2,
                        "Tracker ID": row.get("Tracker ID", ""),
                        "Error": "Remaining Qty. cannot be negative"
                    })
                    continue

                if "billing_done" in tracker_columns:
                    new_values["billing_done"] = "Yes" if balance_to_bill <= 0 and allocated_qty > 0 else "No"

                    if "billing_done" not in insert_columns:
                        insert_columns.append("billing_done")

                if "sent_for_billing" in tracker_columns:
                    invoice_no = clean_text(new_values.get("invoice_no", ""))

                    if billed_qty > 0 or invoice_no != "" or new_values.get("billing_done", "") == "Yes":
                        new_values["sent_for_billing"] = "Yes"
                    elif clean_text(new_values.get("sent_for_billing", "")) == "":
                        new_values["sent_for_billing"] = "No"

                    if "sent_for_billing" not in insert_columns:
                        insert_columns.append("sent_for_billing")

                if "billing_source" in tracker_columns:
                    new_values["billing_source"] = "Full Tracker Correction Upload"
                    insert_columns.append("billing_source")

                if "remark" in tracker_columns and clean_text(new_values.get("remark", "")) == "":
                    new_values["remark"] = "Inserted by full tracker correction upload"

                    if "remark" not in insert_columns:
                        insert_columns.append("remark")

                insert_columns = list(dict.fromkeys(insert_columns))
                value_keys = [f":{col}" for col in insert_columns]

                db_execute(f"""
                INSERT INTO allocation_tracker ({", ".join(insert_columns)})
                VALUES ({", ".join(value_keys)})
                """, new_values, clear_cache=False)

                log_activity("tracker_correction_new", "Inserted new allocation row from correction upload")
                inserted_rows += 1
                continue

            continue

        if action == "delete":
            db_execute(
                "DELETE FROM allocation_tracker WHERE id = :id",
                {"id": tracker_id},
                clear_cache=False
            )
            log_activity("tracker_correction_delete", f"Deleted allocation ID {tracker_id} from correction upload")
            deleted_rows += 1
            continue

        existing = tracker_by_id[tracker_id]
        update_values = {"id": tracker_id}
        set_parts = []

        for db_col, upload_col in replace_map.items():
            if db_col not in tracker_columns or upload_col not in correction_df.columns:
                continue

            raw_value = row.get(upload_col, "")

            if db_col in numeric_columns:
                value = clean_number(raw_value)
            elif db_col in date_columns:
                value, date_error = parse_upload_date(raw_value)

                if date_error:
                    errors.append({
                        "Excel Row": row_no + 2,
                        "Tracker ID": tracker_id,
                        "Error": date_error
                    })
                    value = None
                    break
            elif db_col in yes_no_columns:
                value = normalize_yes_no(raw_value)

                if value == "":
                    errors.append({
                        "Excel Row": row_no + 2,
                        "Tracker ID": tracker_id,
                        "Error": f"{upload_col} must be Yes or No"
                    })
                    break
            else:
                value = clean_text(raw_value)

            update_values[db_col] = value
            set_parts.append(f"{db_col} = :{db_col}")

        else:
            allocated_qty = update_values.get(
                "allocated_qty",
                clean_number(existing.get("allocated_qty", 0))
            )
            billed_qty = update_values.get(
                "billed_qty",
                clean_number(existing.get("billed_qty", 0))
            )
            remaining_qty = update_values.get(
                "balance_to_bill",
                max(allocated_qty - billed_qty, 0)
            )

            if allocated_qty < 0 or billed_qty < 0 or remaining_qty < 0:
                errors.append({
                    "Excel Row": row_no + 2,
                    "Tracker ID": tracker_id,
                    "Error": "Allocated Qty., Billed Qty., and Remaining Qty. cannot be negative"
                })
                continue

            if "balance_to_bill" in tracker_columns and "balance_to_bill" not in update_values:
                update_values["balance_to_bill"] = max(allocated_qty - billed_qty, 0)
                set_parts.append("balance_to_bill = :balance_to_bill")

            if "billing_done" in tracker_columns:
                calculated_balance = update_values.get("balance_to_bill", remaining_qty)
                update_values["billing_done"] = "Yes" if calculated_balance <= 0 and allocated_qty > 0 else "No"

                if "billing_done = :billing_done" not in set_parts:
                    set_parts.append("billing_done = :billing_done")

            if "sent_for_billing" in tracker_columns:
                invoice_no = update_values.get("invoice_no", clean_text(existing.get("invoice_no", "")))

                if billed_qty > 0 or invoice_no != "" or update_values.get("billing_done", "") == "Yes":
                    update_values["sent_for_billing"] = "Yes"

                    if "sent_for_billing = :sent_for_billing" not in set_parts:
                        set_parts.append("sent_for_billing = :sent_for_billing")

            if not set_parts:
                errors.append({
                    "Excel Row": row_no + 2,
                    "Tracker ID": tracker_id,
                    "Error": "No replaceable columns found"
                })
                continue

            db_execute(f"""
            UPDATE allocation_tracker
            SET {", ".join(set_parts)}
            WHERE id = :id
            """, update_values, clear_cache=False)

            log_activity("tracker_correction_replace", f"Replaced allocation ID {tracker_id} from correction upload")
            replaced_rows += 1

    if inserted_rows > 0 or replaced_rows > 0 or deleted_rows > 0:
        st.cache_data.clear()

    return {
        "inserted_rows": inserted_rows,
        "replaced_rows": replaced_rows,
        "deleted_rows": deleted_rows,
        "error_rows": pd.DataFrame(errors)
    }


def render_full_tracker_correction_upload(tracker):
    if st.session_state.role not in ["Admin", "Ops"]:
        return

    st.markdown("---")
    st.subheader("Full Tracker Correction Upload")
    st.caption(
        "Use this to load corrected opening tracker data after reset, "
        "or to fix allocation, billed qty., remaining qty., invoice, or status. "
        "Action can be New, Replace, or Delete."
    )

    correction_template = get_tracker_correction_template_df(tracker)

    with st.expander("Correction upload format"):
        st.dataframe(correction_template.head(25), use_container_width=True)
        st.download_button(
            "Download Full Correction Template",
            data=to_excel(correction_template),
            file_name="full_tracker_correction_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    correction_file = st.file_uploader(
        "Upload Full Tracker Correction File",
        type=["xlsx"],
        key="full_tracker_correction_file"
    )

    if correction_file is not None:
        correction_uploaded_df, correction_read_error = read_uploaded_excel(
            correction_file,
            "Full Tracker Correction file"
        )

        if correction_read_error:
            st.error(correction_read_error)

        else:
            st.write(f"Rows in uploaded file: {len(correction_uploaded_df):,}")
            st.dataframe(correction_uploaded_df.head(25), use_container_width=True)

            confirm_correction = st.checkbox(
                "I understand this will insert, replace, or delete tracker records",
                key="confirm_full_tracker_correction"
            )

            if st.button("Apply Full Tracker Correction Upload"):
                if not confirm_correction:
                    st.error("Please confirm before applying tracker corrections.")

                else:
                    correction_result = apply_tracker_correction_upload(correction_uploaded_df)

                    if not correction_result["error_rows"].empty:
                        st.warning(
                            f"{len(correction_result['error_rows'])} rows could not be corrected. "
                            "Please fix them and upload again."
                        )
                        st.dataframe(correction_result["error_rows"], use_container_width=True)
                        st.download_button(
                            "Download Correction Errors",
                            data=to_excel(correction_result["error_rows"]),
                            file_name="tracker_correction_errors.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                    changed_rows = (
                        correction_result["inserted_rows"] +
                        correction_result["replaced_rows"] +
                        correction_result["deleted_rows"]
                    )

                    if changed_rows > 0:
                        st.success(
                            f"{correction_result['inserted_rows']} rows inserted, "
                            f"{correction_result['replaced_rows']} rows replaced, and "
                            f"{correction_result['deleted_rows']} rows deleted successfully"
                        )
                        st.rerun()


# =====================================================
# FSN BARCODE GENERATION
# =====================================================


@st.cache_data(ttl=3600, show_spinner=False)
def generate_fsn_barcode_bytes(fsn):
    fsn = clean_text(fsn)

    if not fsn:
        return None

    output = BytesIO()
    barcode = Code128(fsn, writer=ImageWriter())
    barcode.write(
        output,
        options={
            "module_width": 0.32,
            "module_height": 18,
            "quiet_zone": 3,
            "write_text": False,
            "dpi": 300,
        }
    )
    output.seek(0)

    return output.getvalue()


def get_fsn_barcode_bytes(fsn, barcode_cache):
    fsn = clean_text(fsn)

    if not fsn:
        return None

    if fsn not in barcode_cache:
        barcode_cache[fsn] = generate_fsn_barcode_bytes(fsn)

    return barcode_cache[fsn]


@st.cache_data(ttl=300, show_spinner=False)
def to_excel(df):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)

    return output.getvalue()


def read_uploaded_excel(uploaded_file, file_label):
    try:
        return pd.read_excel(uploaded_file), ""

    except Exception as exc:
        return pd.DataFrame(), (
            f"Could not read the {file_label}. Please upload a normal Excel workbook "
            f"with at least one visible sheet. Original error: {exc}"
        )


def prepare_direct_billing_df(uploaded_df):
    direct_df = normalize_columns(uploaded_df)

    column_aliases = {
        "Invoice No.": ["Invoice No.", "Invoice No", "Invoice", "Invoice Number", "invoice_no"],
        "PO No.": ["PO No.", "PO No", "PO", "PO Number", "po_no"],
        "FSN": ["FSN", "fsn"],
        "Title": ["Title", "Product Title", "Item Title", "title"],
        "RR Warehouse": ["RR Warehouse", "RR WH", "RR Warehouse Code", "rr_warehouse"],
        "FK Warehouse": ["FK Warehouse", "FK WH", "FK Warehouse Code", "fk_warehouse"],
        "SAP Code": ["SAP Code", "SAP", "SAP SKU", "sap_code"],
        "Qty.": ["Qty.", "Qty", "Quantity", "Billing Qty", "Allocated Qty.", "allocated_qty"],
    }

    rename_map = {}

    for required_col, aliases in column_aliases.items():
        for alias in aliases:
            if alias in direct_df.columns:
                rename_map[alias] = required_col
                break

    direct_df = direct_df.rename(columns=rename_map)

    required_cols = list(column_aliases.keys())
    missing_cols = [col for col in required_cols if col not in direct_df.columns]

    if missing_cols:
        return pd.DataFrame(), missing_cols

    direct_df = direct_df[required_cols].copy()

    for col in [
        "Invoice No.",
        "PO No.",
        "FSN",
        "Title",
        "RR Warehouse",
        "FK Warehouse",
        "SAP Code"
    ]:
        direct_df[col] = direct_df[col].apply(clean_text)

    direct_df["Qty."] = direct_df["Qty."].apply(clean_number)
    direct_df = direct_df[direct_df["FSN"] != ""]

    return direct_df, []


# =====================================================
# BILLING SUMMARY WITH EXACT BARCODES
# =====================================================

@st.cache_data(ttl=300, show_spinner=False)
def to_excel_billing_with_exact_barcodes(billing_df):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Billing Summary"
    fsn_barcode_cache = {}

    headers = [
        "Invoice No.",
        "PO No.",
        "FSN",
        "FSN Barcode",
        "Title",
        "RR Warehouse",
        "FK Warehouse",
        "SAP Code",
        "Qty."
    ]

    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(
            bold=True,
            color="FFFFFF"
        )

        ws.cell(row=1, column=col).fill = PatternFill(
            "solid",
            fgColor="1F4E78"
        )

        ws.cell(row=1, column=col).alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 12

    temp_files = []

    for idx, row in billing_df.iterrows():
        excel_row = idx + 2
        fsn = clean_text(row["FSN"])

        ws.cell(excel_row, 1).value = row["Invoice No."]
        ws.cell(excel_row, 2).value = row["PO No."]
        ws.cell(excel_row, 3).value = fsn
        ws.cell(excel_row, 5).value = row["Title"]
        ws.cell(excel_row, 6).value = row["RR Warehouse"]
        ws.cell(excel_row, 7).value = row["FK Warehouse"]
        ws.cell(excel_row, 8).value = row["SAP Code"]
        ws.cell(excel_row, 9).value = row["Qty."]

        ws.row_dimensions[excel_row].height = 115

        barcode_bytes = get_fsn_barcode_bytes(fsn, fsn_barcode_cache)

        if barcode_bytes:
            temp_file = tempfile.NamedTemporaryFile(
                delete=False,
                suffix=".png"
            )

            temp_file.write(barcode_bytes)
            temp_file.close()
            temp_files.append(temp_file.name)

            xl_img = XLImage(temp_file.name)
            xl_img.width = 340
            xl_img.height = 120

            ws.add_image(
                xl_img,
                f"D{excel_row}"
            )

        else:
            ws.cell(excel_row, 4).value = "Barcode Not Found"

        for col in range(1, 10):
            ws.cell(excel_row, col).alignment = Alignment(
                vertical="center",
                horizontal="center",
                wrap_text=True
            )

    wb.save(output)

    for file in temp_files:
        try:
            os.remove(file)
        except:
            pass

    output.seek(0)

    return output.getvalue()


# =====================================================
# STICKER PDF GENERATOR
# =====================================================

@st.cache_data(ttl=300, show_spinner=False)
def generate_sticker_pdf(billing_df):
    pdf_buffer = BytesIO()
    fsn_barcode_cache = {}

    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=A4,
        rightMargin=8,
        leftMargin=8,
        topMargin=8,
        bottomMargin=8
    )

    elements = []
    temp_files = []

    for _, row in billing_df.iterrows():
        fsn = clean_text(row["FSN"])
        title = str(row["Title"])
        invoice_text = f"Invoice No. - {row['Invoice No.']}"
        po_text = f"PO No. - {row['PO No.']}"

        sticker_data = []

        for i in range(8):
            barcode_bytes = get_fsn_barcode_bytes(fsn, fsn_barcode_cache)

            if barcode_bytes:
                temp_file = tempfile.NamedTemporaryFile(
                    delete=False,
                    suffix=".png"
                )

                temp_file.write(barcode_bytes)
                temp_file.close()
                temp_files.append(temp_file.name)

                barcode_img = Image(
                    temp_file.name,
                    width=68 * mm,
                    height=22 * mm
                )

            else:
                barcode_img = Table(
                    [["Barcode Not Found"]],
                    colWidths=[75 * mm]
                )

            sticker_inner = Table(
                [
                    [title],
                    [barcode_img],
                    [fsn],
                    [invoice_text],
                    [po_text]
                ],
                colWidths=[82 * mm],
                rowHeights=[
                    10 * mm,
                    25 * mm,
                    7 * mm,
                    7 * mm,
                    7 * mm
                ]
            )

            sticker_inner.setStyle(TableStyle([
                ("BOX", (0, 0), (-1, -1), 1, colors.black),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTNAME", (0, 0), (0, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (0, 0), 7),
                ("FONTSIZE", (0, 2), (0, 4), 7),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]))

            sticker_data.append(sticker_inner)

        rows = []

        for i in range(0, len(sticker_data), 2):
            rows.append(sticker_data[i:i + 2])

        final_table = Table(
            rows,
            colWidths=[90 * mm, 90 * mm],
            rowHeights=[65 * mm] * 4
        )

        final_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))

        elements.append(final_table)
        elements.append(PageBreak())

    if elements and isinstance(elements[-1], PageBreak):
        elements.pop()

    doc.build(elements)

    for file in temp_files:
        try:
            os.remove(file)
        except:
            pass

    pdf_buffer.seek(0)

    return pdf_buffer.getvalue()


# =====================================================
# ALLOCATION ENGINE
# =====================================================

def run_allocation(pending_df, stock_df):
    pending_df = standardize_pending_file(pending_df)
    stock_df = standardize_stock_file(stock_df)

    required_pending = [
        "PO No.", "FSN", "Title", "RR Warehouse", "FK Warehouse", "Pending Qty."
    ]

    required_stock = [
        "FSN", "RR Warehouse", "SAP Code", "Stock"
    ]

    missing_pending = [c for c in required_pending if c not in pending_df.columns]
    missing_stock = [c for c in required_stock if c not in stock_df.columns]

    if missing_pending:
        st.error(f"Missing columns in Pending Qty file: {missing_pending}")
        st.write("Pending file columns found:", list(pending_df.columns))
        return pd.DataFrame()

    if missing_stock:
        st.error(f"Missing columns in Stock file: {missing_stock}")
        st.write("Stock file columns found:", list(stock_df.columns))
        return pd.DataFrame()

    result = []

    for col in ["PO No.", "Order ID", "Buyer Code", "FSN", "Title", "RR Warehouse", "FK Warehouse"]:
        if col not in pending_df.columns:
            pending_df[col] = ""
        pending_df[col] = pending_df[col].apply(clean_text)

    pending_df["Pending Qty."] = pending_df["Pending Qty."].apply(clean_number)

    if "Pending Amount" not in pending_df.columns:
        pending_df["Pending Amount"] = 0

    pending_df["Pending Amount"] = pending_df["Pending Amount"].apply(clean_number)

    existing_po_alloc = get_existing_po_allocation_qty()

    if not existing_po_alloc.empty:
        existing_po_alloc["po_no"] = existing_po_alloc["po_no"].apply(clean_text)
        existing_po_alloc["fsn"] = existing_po_alloc["fsn"].apply(clean_text)
        existing_po_alloc["rr_warehouse"] = existing_po_alloc["rr_warehouse"].apply(clean_text)
        existing_po_alloc["fk_warehouse"] = existing_po_alloc["fk_warehouse"].apply(clean_text)

        pending_df = pending_df.merge(
            existing_po_alloc,
            how="left",
            left_on=["PO No.", "FSN", "RR Warehouse", "FK Warehouse"],
            right_on=["po_no", "fsn", "rr_warehouse", "fk_warehouse"]
        )

        pending_df["already_allocated_qty"] = pending_df["already_allocated_qty"].fillna(0)

    else:
        pending_df["already_allocated_qty"] = 0

    pending_df["Allocatable Pending Qty."] = (
        pending_df["Pending Qty."] - pending_df["already_allocated_qty"]
    )

    pending_df["Allocatable Pending Qty."] = pending_df["Allocatable Pending Qty."].apply(
        lambda x: max(x, 0)
    )

    for col in ["FSN", "RR Warehouse", "SAP Code"]:
        stock_df[col] = stock_df[col].apply(clean_text)

    stock_df["Stock"] = stock_df["Stock"].apply(clean_number)

    open_alloc = get_open_allocation_qty()

    if not open_alloc.empty:
        open_alloc["fsn"] = open_alloc["fsn"].apply(clean_text)
        open_alloc["rr_warehouse"] = open_alloc["rr_warehouse"].apply(clean_text)
        open_alloc["sap_code"] = open_alloc["sap_code"].apply(clean_text)

        stock_df = stock_df.merge(
            open_alloc,
            how="left",
            left_on=["FSN", "RR Warehouse", "SAP Code"],
            right_on=["fsn", "rr_warehouse", "sap_code"]
        )

        stock_df["open_alloc_qty"] = stock_df["open_alloc_qty"].fillna(0)

    else:
        stock_df["open_alloc_qty"] = 0

    stock_df["usable_stock"] = stock_df["Stock"] - stock_df["open_alloc_qty"]
    stock_df["usable_stock"] = stock_df["usable_stock"].apply(lambda x: max(x, 0))
    appointment_df = get_appointment_df()

    for _, p in pending_df.iterrows():
        po = clean_text(p["PO No."])
        order_id = clean_text(p["Order ID"])
        buyer_code = clean_text(p["Buyer Code"])
        fsn = clean_text(p["FSN"])
        title = clean_text(p["Title"])
        rr = clean_text(p["RR Warehouse"])
        fk = clean_text(p["FK Warehouse"])
        uploaded_pending_qty = clean_number(p["Pending Qty."])
        pending_amount = clean_number(p["Pending Amount"])
        already_allocated_qty = clean_number(p["already_allocated_qty"])
        pending_qty = clean_number(p["Allocatable Pending Qty."])
        pending_unit_amount = pending_amount / uploaded_pending_qty if uploaded_pending_qty > 0 else 0

        remaining = pending_qty
        allocated_anything = False
        pending_shown_for_demand = False
        allocation_row_indexes = []
        appointment_matches = find_appointment_matches(appointment_df, po, fsn, rr, fk)

        if not appointment_matches.empty:
            allocation_contexts = []

            for appt_idx, appt in appointment_matches.iterrows():
                appointment_balance = clean_number(appt["appointment_balance_qty"])

                if appointment_balance <= 0:
                    continue

                allocation_contexts.append({
                    "appointment_index": appt_idx,
                    "appointment_no": clean_text(appt["appointment_no"]),
                    "appointment_date": clean_text(appt["appointment_date"]),
                    "qty": appointment_balance,
                })

        else:
            allocation_contexts = [{
                "appointment_index": None,
                "appointment_no": "",
                "appointment_date": "",
                "qty": pending_qty,
            }]

        if pending_qty <= 0:
            result.append({
                "PO No.": po,
                "Order ID": order_id,
                "Buyer Code": buyer_code,
                "FSN": fsn,
                "Title": title,
                "RR Warehouse": rr,
                "FK Warehouse": fk,
                "SAP Code": "",
                "Appointment No.": "",
                "Appointment Date": "",
                "Pending Qty.": pending_qty,
                "Pending Amount": 0,
                "Allocated Qty.": 0,
                "Balance Pending Qty.": 0,
                "Current Stock": 0,
                "Open Allocation Qty": 0,
                "Usable Stock Before": 0,
                "Usable Stock After": 0,
                "Status": "Already Allocated"
            })
            continue

        matching = stock_df[
            (stock_df["FSN"] == fsn) &
            (stock_df["RR Warehouse"] == rr) &
            (stock_df["usable_stock"] > 0)
        ]

        for appointment_context in allocation_contexts:
            if remaining <= 0:
                break

            appointment_remaining = min(
                remaining,
                clean_number(appointment_context["qty"])
            )

            if appointment_remaining <= 0:
                continue

            for idx, s in matching.iterrows():
                if remaining <= 0 or appointment_remaining <= 0:
                    break

                usable = clean_number(stock_df.loc[idx, "usable_stock"])
                alloc = min(usable, remaining, appointment_remaining)

                if alloc > 0:
                    row_pending_qty = pending_qty if not pending_shown_for_demand else 0
                    row_pending_amount = pending_qty * pending_unit_amount if not pending_shown_for_demand else 0

                    allocated_anything = True
                    pending_shown_for_demand = True
                    stock_df.loc[idx, "usable_stock"] = usable - alloc
                    remaining -= alloc
                    appointment_remaining -= alloc

                    appt_idx = appointment_context["appointment_index"]

                    if appt_idx is not None:
                        appointment_df.loc[appt_idx, "appointment_balance_qty"] = (
                            clean_number(appointment_df.loc[appt_idx, "appointment_balance_qty"]) - alloc
                        )

                    allocation_row_indexes.append(len(result))

                    result.append({
                        "PO No.": po,
                        "Order ID": order_id,
                        "Buyer Code": buyer_code,
                        "FSN": fsn,
                        "Title": title,
                        "RR Warehouse": rr,
                        "FK Warehouse": fk,
                        "SAP Code": s["SAP Code"],
                        "Appointment No.": appointment_context["appointment_no"],
                        "Appointment Date": appointment_context["appointment_date"],
                        "Pending Qty.": row_pending_qty,
                        "Pending Amount": row_pending_amount,
                        "Allocated Qty.": alloc,
                        "Balance Pending Qty.": 0,
                        "Current Stock": s["Stock"],
                        "Open Allocation Qty": s["open_alloc_qty"],
                        "Usable Stock Before": usable,
                        "Usable Stock After": usable - alloc,
                        "Status": "Allocated"
                    })

        if allocation_row_indexes:
            result[allocation_row_indexes[-1]]["Balance Pending Qty."] = remaining

        if not allocated_anything:
            result.append({
                "PO No.": po,
                "Order ID": order_id,
                "Buyer Code": buyer_code,
                "FSN": fsn,
                "Title": title,
                "RR Warehouse": rr,
                "FK Warehouse": fk,
                "SAP Code": "",
                "Appointment No.": "",
                "Appointment Date": "",
                "Pending Qty.": pending_qty,
                "Pending Amount": pending_qty * pending_unit_amount,
                "Allocated Qty.": 0,
                "Balance Pending Qty.": remaining,
                "Current Stock": 0,
                "Open Allocation Qty": 0,
                "Usable Stock Before": 0,
                "Usable Stock After": 0,
                "Status": "No Stock"
            })

    return pd.DataFrame(result)


# =====================================================
# SAVE ALLOCATION TO TRACKER
# =====================================================

def save_allocation(df):
    saved_count = 0
    tracker_columns = get_allocation_tracker_columns()
    has_order_id = "order_id" in tracker_columns
    has_buyer_code = "buyer_code" in tracker_columns
    has_pending_amount = "pending_amount" in tracker_columns
    has_billed_qty = "billed_qty" in tracker_columns
    has_balance_to_bill = "balance_to_bill" in tracker_columns
    has_billing_source = "billing_source" in tracker_columns
    has_appointment_no = "appointment_no" in tracker_columns
    has_appointment_date = "appointment_date" in tracker_columns

    for _, r in df.iterrows():
        if clean_number(r["Allocated Qty."]) <= 0:
            continue

        insert_columns = [
            "allocation_date",
            "po_no",
            "fsn",
            "title",
            "rr_warehouse",
            "fk_warehouse",
            "sap_code",
            "allocated_qty",
            "sent_for_billing",
            "billing_done",
            "remark"
        ]

        value_keys = [
            ":allocation_date",
            ":po_no",
            ":fsn",
            ":title",
            ":rr_warehouse",
            ":fk_warehouse",
            ":sap_code",
            ":allocated_qty",
            ":sent_for_billing",
            ":billing_done",
            ":remark"
        ]

        params = {
            "allocation_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "po_no": r["PO No."],
            "fsn": r["FSN"],
            "title": r["Title"],
            "rr_warehouse": r["RR Warehouse"],
            "fk_warehouse": r["FK Warehouse"],
            "sap_code": r["SAP Code"],
            "allocated_qty": clean_number(r["Allocated Qty."]),
            "sent_for_billing": "No",
            "billing_done": "No",
            "remark": "Fresh Allocation"
        }

        if has_order_id:
            insert_columns.insert(2, "order_id")
            value_keys.insert(2, ":order_id")
            params["order_id"] = r.get("Order ID", "")

        if has_buyer_code:
            insert_at = 3 if has_order_id else 2
            insert_columns.insert(insert_at, "buyer_code")
            value_keys.insert(insert_at, ":buyer_code")
            params["buyer_code"] = r.get("Buyer Code", "")

        if has_pending_amount:
            insert_at = 9
            if has_order_id:
                insert_at += 1
            if has_buyer_code:
                insert_at += 1
            insert_columns.insert(insert_at, "pending_amount")
            value_keys.insert(insert_at, ":pending_amount")
            params["pending_amount"] = clean_number(r.get("Pending Amount", 0))

        if has_appointment_no:
            insert_columns.insert(-3, "appointment_no")
            value_keys.insert(-3, ":appointment_no")
            params["appointment_no"] = r.get("Appointment No.", "")

        if has_appointment_date:
            insert_columns.insert(-3, "appointment_date")
            value_keys.insert(-3, ":appointment_date")
            params["appointment_date"] = r.get("Appointment Date", "")

        if has_billed_qty:
            insert_columns.insert(-3, "billed_qty")
            value_keys.insert(-3, ":billed_qty")
            params["billed_qty"] = 0

        if has_balance_to_bill:
            insert_columns.insert(-3, "balance_to_bill")
            value_keys.insert(-3, ":balance_to_bill")
            params["balance_to_bill"] = clean_number(r["Allocated Qty."])

        if has_billing_source:
            insert_columns.insert(-1, "billing_source")
            value_keys.insert(-1, ":billing_source")
            params["billing_source"] = "Allocation"

        db_execute(f"""
        INSERT INTO allocation_tracker (
            {", ".join(insert_columns)}
        )
        VALUES (
            {", ".join(value_keys)}
        )
        """, params, clear_cache=False)

        saved_count += 1

    if saved_count > 0:
        st.cache_data.clear()

    return saved_count


def reset_tracker_data():
    db_execute("DELETE FROM allocation_tracker")


# =====================================================
# SIDEBAR
# =====================================================

menu = st.sidebar.radio(
    "Select Screen",
    get_allowed_screens(st.session_state.role)
)

st.sidebar.markdown("---")
st.sidebar.write(f"Logged in as: {st.session_state.username}")
st.sidebar.write(f"Role: {st.session_state.role}")

if st.sidebar.button("Logout"):
    log_activity("logout", "User logged out")
    st.session_state.clear()
    st.rerun()


# =====================================================
# DASHBOARD
# =====================================================

if menu == "Dashboard Summary":
    render_page_header(
        "Dashboard Summary",
        "A quick view of allocation, billing, and open stock movement."
    )

    tracker = get_tracker_df()
    open_alloc = get_open_allocation_qty()

    total_alloc = tracker["allocated_qty"].sum() if not tracker.empty else 0
    sent_qty = tracker[tracker["sent_for_billing"] == "Yes"]["allocated_qty"].sum() if not tracker.empty else 0
    if not tracker.empty and "billed_qty" in tracker.columns:
        billed_qty = tracker["billed_qty"].fillna(0).sum()
    else:
        billed_qty = tracker[tracker["billing_done"] == "Yes"]["allocated_qty"].sum() if not tracker.empty else 0
    open_qty = open_alloc["open_alloc_qty"].sum() if not open_alloc.empty else 0

    c1, c2, c3, c4 = st.columns(4)

    c1.markdown(f"""
    <div class="metric-card blue">
        <div class="metric-title">Total Allocated</div>
        <div class="metric-value">{total_alloc:,.0f}</div>
    </div>
    """, unsafe_allow_html=True)

    c2.markdown(f"""
    <div class="metric-card orange">
        <div class="metric-title">Sent for Billing</div>
        <div class="metric-value">{sent_qty:,.0f}</div>
    </div>
    """, unsafe_allow_html=True)

    c3.markdown(f"""
    <div class="metric-card green">
        <div class="metric-title">Billed Qty</div>
        <div class="metric-value">{billed_qty:,.0f}</div>
    </div>
    """, unsafe_allow_html=True)

    c4.markdown(f"""
    <div class="metric-card red">
        <div class="metric-title">Open Allocation</div>
        <div class="metric-value">{open_qty:,.0f}</div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("---")

    if not tracker.empty:
        st.subheader("Date-wise Billing Summary")

        summary_c1, summary_c2 = st.columns(2)

        with summary_c1:
            sent_summary = tracker.copy()
            sent_summary["sent_date"] = pd.to_datetime(
                sent_summary["sent_date"],
                errors="coerce"
            ).dt.date
            sent_summary = sent_summary[
                (sent_summary["sent_for_billing"] == "Yes") &
                (sent_summary["sent_date"].notna())
            ]

            if sent_summary.empty:
                st.info("No date-wise sent for billing data yet.")
            else:
                sent_summary = sent_summary.groupby(
                    "sent_date",
                    as_index=False
                )["allocated_qty"].sum().rename(columns={
                    "sent_date": "Sent Date",
                    "allocated_qty": "Sent Qty"
                })
                st.markdown("#### Sent for Billing")
                st.dataframe(sent_summary, use_container_width=True)

        with summary_c2:
            billed_summary = tracker.copy()
            billed_summary["billing_date"] = pd.to_datetime(
                billed_summary["billing_date"],
                errors="coerce"
            ).dt.date
            billed_summary = billed_summary[
                (billed_summary["billing_done"] == "Yes") &
                (billed_summary["billing_date"].notna())
            ]

            if billed_summary.empty:
                st.info("No date-wise billing done data yet.")
            else:
                billed_qty_column = "billed_qty" if "billed_qty" in billed_summary.columns else "allocated_qty"
                billed_summary[billed_qty_column] = billed_summary[billed_qty_column].fillna(0)
                billed_summary = billed_summary.groupby(
                    "billing_date",
                    as_index=False
                )[billed_qty_column].sum().rename(columns={
                    "billing_date": "Billing Date",
                    billed_qty_column: "Billed Qty"
                })
                st.markdown("#### Billing Done")
                st.dataframe(billed_summary, use_container_width=True)

        st.markdown("---")
        st.subheader("Recent Allocation Records")
        recent_tracker = tracker.head(500)
        st.caption(f"Showing latest {len(recent_tracker):,} of {len(tracker):,} records.")
        st.dataframe(recent_tracker, use_container_width=True)
    else:
        st.info("No allocation records yet.")

    if st.session_state.role == "Admin":
        st.markdown("---")
        st.subheader("System Reset")

        with st.expander("Reset Tracker Data"):
            st.markdown("""
<div style="
background: linear-gradient(135deg,#fee2e2,#fecaca);
border-left: 6px solid #dc2626;
padding: 16px;
border-radius: 14px;
font-weight: 700;
color: #7f1d1d;
margin-top: 10px;
margin-bottom: 10px;
">
This will permanently delete all saved allocation tracker data.
</div>
""", unsafe_allow_html=True)
            confirm_reset = st.checkbox(
                "I understand and want to reset tracker data"
            )

            if confirm_reset:
                if st.button("Reset All Tracker Data"):
                    reset_tracker_data()
                    log_activity("reset_tracker", "All tracker data reset")
                    st.success("Tracker data reset successfully")
                    st.rerun()


# =====================================================
# MARKETING DASHBOARD
# =====================================================

elif menu == "Marketing Dashboard":
    show_marketing_dashboard(
        engine,
        db_read,
        db_execute,
        db_execute_many,
        clean_text,
        clean_number,
    )
    st.stop()

    render_page_header(
        "Marketing Dashboard",
        "A product and customer view of allocation, billing, and pending billing performance."
    )

    tracker = get_tracker_df()

    if tracker.empty:
        st.info("No tracker data available yet.")

    else:
        marketing_df = tracker.copy()

        for col in ["allocated_qty", "billed_qty", "balance_to_bill", "pending_amount"]:
            if col not in marketing_df.columns:
                marketing_df[col] = 0
            marketing_df[col] = marketing_df[col].apply(clean_number)

        for col in ["buyer_code", "fk_warehouse", "rr_warehouse", "fsn", "title", "sap_code"]:
            if col not in marketing_df.columns:
                marketing_df[col] = ""
            marketing_df[col] = marketing_df[col].apply(clean_text)

        marketing_df["Billing Status"] = marketing_df.apply(
            lambda row: (
                "Fully Billed"
                if clean_number(row.get("balance_to_bill", 0)) <= 0 and clean_number(row.get("allocated_qty", 0)) > 0
                else "Partially Billed"
                if clean_number(row.get("billed_qty", 0)) > 0
                else "Pending Billing"
            ),
            axis=1
        )

        total_allocated = marketing_df["allocated_qty"].sum()
        total_billed = marketing_df["billed_qty"].sum()
        total_balance = marketing_df["balance_to_bill"].sum()
        billing_rate = (total_billed / total_allocated * 100) if total_allocated else 0

        c1, c2, c3, c4 = st.columns(4)

        c1.markdown(f"""
        <div class="metric-card blue">
            <div class="metric-title">Allocated Qty</div>
            <div class="metric-value">{total_allocated:,.0f}</div>
        </div>
        """, unsafe_allow_html=True)

        c2.markdown(f"""
        <div class="metric-card green">
            <div class="metric-title">Billed Qty</div>
            <div class="metric-value">{total_billed:,.0f}</div>
        </div>
        """, unsafe_allow_html=True)

        c3.markdown(f"""
        <div class="metric-card orange">
            <div class="metric-title">Balance To Bill</div>
            <div class="metric-value">{total_balance:,.0f}</div>
        </div>
        """, unsafe_allow_html=True)

        c4.markdown(f"""
        <div class="metric-card red">
            <div class="metric-title">Billing %</div>
            <div class="metric-value">{billing_rate:,.1f}%</div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("---")

        filter_c1, filter_c2, filter_c3 = st.columns(3)

        with filter_c1:
            buyer_options = sorted([x for x in marketing_df["buyer_code"].unique() if x != ""])
            selected_buyer = st.multiselect("Buyer Code", buyer_options)

        with filter_c2:
            fk_options = sorted([x for x in marketing_df["fk_warehouse"].unique() if x != ""])
            selected_fk = st.multiselect("FK Warehouse", fk_options)

        with filter_c3:
            status_options = sorted(marketing_df["Billing Status"].unique())
            selected_status = st.multiselect("Billing Status", status_options)

        filtered_df = marketing_df.copy()

        if selected_buyer:
            filtered_df = filtered_df[filtered_df["buyer_code"].isin(selected_buyer)]

        if selected_fk:
            filtered_df = filtered_df[filtered_df["fk_warehouse"].isin(selected_fk)]

        if selected_status:
            filtered_df = filtered_df[filtered_df["Billing Status"].isin(selected_status)]

        st.caption(f"Showing {len(filtered_df):,} tracker rows after filters.")

        tab1, tab2, tab3, tab4 = st.tabs([
            "Buyer View",
            "Warehouse View",
            "Product View",
            "Pending Billing Detail"
        ])

        def build_summary(group_cols):
            if filtered_df.empty:
                return pd.DataFrame()

            summary_df = filtered_df.groupby(
                group_cols,
                dropna=False,
                as_index=False
            ).agg(
                Allocated_Qty=("allocated_qty", "sum"),
                Billed_Qty=("billed_qty", "sum"),
                Balance_To_Bill=("balance_to_bill", "sum"),
                Pending_Amount=("pending_amount", "sum"),
                Tracker_Rows=("id", "count")
            )

            summary_df["Billing %"] = summary_df.apply(
                lambda row: (
                    clean_number(row["Billed_Qty"]) /
                    clean_number(row["Allocated_Qty"]) * 100
                ) if clean_number(row["Allocated_Qty"]) else 0,
                axis=1
            )

            return summary_df.sort_values(
                by=["Balance_To_Bill", "Allocated_Qty"],
                ascending=[False, False]
            )

        with tab1:
            st.subheader("Buyer Code Summary")
            buyer_summary = build_summary(["buyer_code"])
            st.dataframe(buyer_summary, use_container_width=True)

        with tab2:
            st.subheader("Warehouse Summary")
            warehouse_summary = build_summary(["fk_warehouse", "rr_warehouse"])
            st.dataframe(warehouse_summary, use_container_width=True)

        with tab3:
            st.subheader("Product Summary")
            product_summary = build_summary(["fsn", "title", "sap_code"])
            st.dataframe(product_summary.head(500), use_container_width=True)

        with tab4:
            st.subheader("Pending Billing Detail")
            pending_detail = filtered_df[
                filtered_df["balance_to_bill"] > 0
            ].copy()

            detail_cols = [
                "po_no",
                "order_id",
                "buyer_code",
                "fsn",
                "title",
                "rr_warehouse",
                "fk_warehouse",
                "sap_code",
                "allocated_qty",
                "billed_qty",
                "balance_to_bill",
                "sent_date",
                "remark",
            ]
            detail_cols = [col for col in detail_cols if col in pending_detail.columns]

            st.dataframe(
                pending_detail[detail_cols].head(500),
                use_container_width=True
            )

            if not pending_detail.empty:
                st.download_button(
                    "Download Pending Billing Detail",
                    data=to_excel(pending_detail[detail_cols]),
                    file_name="marketing_pending_billing_detail.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )


# =====================================================
# UPLOAD & ALLOCATE
# =====================================================

elif menu == "Upload & Allocate":
    render_page_header(
        "Upload & Allocate",
        "Upload pending PO demand and stock sheets, then generate fresh allocation."
    )

    st.markdown("#### Required Upload Formats")

    sample1, sample2 = st.columns(2)

    with sample1:
        st.markdown("#### Pending Qty File Format")
        pending_sample = pd.DataFrame({
            "PO No.": ["PO123"],
            "Order ID": ["SO-123"],
            "Buyer Code": ["CUS001"],
            "FSN": ["FSN001"],
            "Title": ["Fan"],
            "RR Warehouse": ["WH1"],
            "FK Warehouse": ["FK1"],
            "Pending Qty.": [100],
            "Pending Amount": [25000]
        })
        st.dataframe(pending_sample, use_container_width=True)
        st.caption(
            "System pending files are also accepted. The app maps PONo, Order Id, "
            "Buyer Code, Item Description, Site Id, FK FC, Pending Qty and Pending Amt. "
            "Both Hold and Pending status rows are included."
        )

    with sample2:
        st.markdown("#### Stock File Format")
        stock_sample = pd.DataFrame({
            "FSN": ["FSN001"],
            "RR Warehouse": ["WH1"],
            "SAP Code": ["SAP001"],
            "Stock": [200]
        })
        st.dataframe(stock_sample, use_container_width=True)
        st.caption(
            "System stock files are also accepted. The app maps Item Code to SAP Code, "
            "Available Qty to Stock, Site to RR Warehouse, and filters CustomLocation "
            "to FGI/ECOM plus the approved SITE list."
        )

    st.markdown("---")
    st.subheader("Appointment Upload")
    st.caption(
        "Upload Flipkart appointment quantity before running allocation. "
        "When an appointment exists for FSN/FK warehouse, allocation is capped by appointment balance."
    )

    appointment_sample = pd.DataFrame({
        "Appointment No.": ["APT123"],
        "Appointment Date": ["2026-05-25"],
        "PO No.": ["PO123"],
        "FSN": ["FSN001"],
        "FK Warehouse": ["FK1"],
        "Appointment Qty.": [50],
        "Remark": ["Scheduled delivery appointment"],
    })

    with st.expander("Appointment upload format"):
        st.dataframe(appointment_sample, use_container_width=True)
        st.download_button(
            "Download Appointment Upload Template",
            data=to_excel(appointment_sample),
            file_name="appointment_upload_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    appointment_file = st.file_uploader(
        "Upload Appointment File",
        type=["xlsx"],
        key="appointment_file"
    )

    if appointment_file is not None:
        appointment_uploaded_df, appointment_read_error = read_uploaded_excel(
            appointment_file,
            "Appointment file"
        )

        if appointment_read_error:
            st.error(appointment_read_error)

        else:
            appointment_df, missing_appointment_cols = prepare_appointment_upload(
                appointment_uploaded_df
            )

            if missing_appointment_cols:
                st.error(f"Missing columns in Appointment file: {missing_appointment_cols}")

            elif appointment_df.empty:
                st.warning("No valid appointment rows found.")

            else:
                st.dataframe(appointment_df, use_container_width=True)

                if st.button("Save Appointment Upload"):
                    saved_appointments = save_appointment_upload(appointment_df)
                    log_activity("appointment_upload", f"{saved_appointments} appointment rows saved")
                    st.success(f"{saved_appointments} appointment rows saved successfully")
                    st.rerun()

    active_appointments = get_appointment_df()

    if not active_appointments.empty:
        with st.expander("Active appointment balance"):
            appointment_display = active_appointments.drop(
                columns=["rr_warehouse"],
                errors="ignore"
            )
            st.dataframe(appointment_display, use_container_width=True)

    st.markdown("---")
    st.subheader("Upload Files")

    c1, c2 = st.columns(2)

    with c1:
        stock_file = st.file_uploader(
            "Upload Stock File",
            type=["xlsx"],
            key="stock_file"
        )

    with c2:
        pending_file = st.file_uploader(
            "Upload Pending Qty File",
            type=["xlsx"],
            key="pending_file"
        )

    if stock_file is not None:
        st.session_state.stock_df = standardize_stock_file(
            pd.read_excel(stock_file),
            show_messages=True
        )

    if pending_file is not None:
        st.session_state.pending_df = standardize_pending_file(
            pd.read_excel(pending_file),
            show_messages=True
        )

    if st.session_state.stock_df is not None and st.session_state.pending_df is not None:
        st.markdown("""
        <div class="success-box">
        Files uploaded successfully
        </div>
        """, unsafe_allow_html=True)

        p1, p2 = st.columns(2)

        with p1:
            st.subheader("Stock Preview")
            st.write(f"Rows uploaded: {len(st.session_state.stock_df)}")
            st.dataframe(st.session_state.stock_df, use_container_width=True)

        with p2:
            st.subheader("Pending Qty Preview")
            st.write(f"Rows uploaded: {len(st.session_state.pending_df)}")
            st.dataframe(st.session_state.pending_df, use_container_width=True)

        st.markdown("---")

        if st.button("Run Allocation"):
            allocation = run_allocation(
                st.session_state.pending_df,
                st.session_state.stock_df
            )
            st.session_state.allocation_df = allocation

        if st.session_state.allocation_df is not None:
            st.subheader("Allocation Output")
            st.dataframe(st.session_state.allocation_df, use_container_width=True)

            b1, b2, b3 = st.columns(3)

            with b1:
                if st.button("Save Allocation to Tracker"):
                    saved = save_allocation(st.session_state.allocation_df)
                    st.success(f"{saved} allocation rows saved successfully")

            with b2:
                st.download_button(
                    "Download Allocation Output",
                    data=to_excel(st.session_state.allocation_df),
                    file_name="allocation_output.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            with b3:
                if st.button("Clear Uploaded Files"):
                    st.session_state.stock_df = None
                    st.session_state.pending_df = None
                    st.session_state.allocation_df = None
                    st.rerun()


# =====================================================
# ALLOCATION TRACKER
# =====================================================

elif menu == "Allocation Tracker":
    render_page_header(
        "Allocation Tracker",
        "Update billing status, invoice details, dates, and remarks for saved allocations."
    )

    tracker = get_tracker_df()

    if tracker.empty:
        st.info("No allocation records found.")
        render_full_tracker_correction_upload(tracker)

    else:
        tracker_task = st.selectbox(
            "Select Tracker Task",
            [
                "View / Manual Edit Table",
                "Sent For Billing Download",
                "Billing Update Upload",
                "Sales Billing Auto Update Upload",
                "Full Tracker Correction Upload",
                "Emergency Manual Allocation Upload",
                "Delete Allocation Upload",
            ]
        )
        st.caption("Select 'View / Manual Edit Table' when you want to see or manually update tracker rows.")

        if tracker_task == "Sent For Billing Download":
            sent_for_billing_export = get_sent_for_billing_download_df(tracker)

            st.markdown("---")
            st.subheader("Sent For Billing Download")

            if sent_for_billing_export.empty:
                st.info("No rows are currently marked as Sent for Billing.")
            else:
                st.download_button(
                    "Download Sent For Billing Rows",
                    data=to_excel(sent_for_billing_export),
                    file_name="sent_for_billing_rows.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        if tracker_task == "Billing Update Upload":
            st.markdown("---")
            st.subheader("Billing Update Upload")
            st.caption(
                "Download this Excel, fill Invoice No., Billing Date as YYYY-MM-DD, "
                "Sent For Billing as Yes/No, Billed Qty., Billing Done as Yes/No, and upload it back. "
                "Tracker ID must not be changed."
            )

            billing_update_template = get_billing_update_template_df(tracker)

            st.download_button(
                "Download Billing Update Template",
                data=to_excel(billing_update_template),
                file_name="billing_update_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            billing_update_file = st.file_uploader(
                "Upload Completed Billing Update File",
                type=["xlsx"],
                key="billing_update_file"
            )

            if billing_update_file is not None:
                uploaded_df, read_error = read_uploaded_excel(
                    billing_update_file,
                    "Billing Update file"
                )

                if read_error:
                    st.error(read_error)

                else:
                    st.write(f"Rows in uploaded file: {len(uploaded_df):,}")
                    st.dataframe(uploaded_df.head(25), use_container_width=True)

                    if st.button("Apply Billing Update Upload"):
                        update_result = apply_billing_update_upload(uploaded_df)

                        if not update_result["error_rows"].empty:
                            st.warning(
                                f"{len(update_result['error_rows'])} rows could not be updated. "
                                "Please fix them and upload again."
                            )
                            st.dataframe(update_result["error_rows"], use_container_width=True)
                            st.download_button(
                                "Download Billing Update Errors",
                                data=to_excel(update_result["error_rows"]),
                                file_name="billing_update_errors.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                        if update_result["updated_rows"] > 0:
                            st.session_state.latest_billing_update_ids = update_result["updated_ids"]
                            log_activity(
                                "billing_update_upload",
                                f"{update_result['updated_rows']} tracker rows updated from upload"
                            )
                            st.success(f"{update_result['updated_rows']} tracker rows updated successfully")
                            st.rerun()

        if tracker_task == "Sales Billing Auto Update Upload":
            st.markdown("---")
            st.subheader("Sales Billing Auto Update Upload")
            st.caption(
                "Upload the system sales file to compare invoice billing against tracker billed qty. "
                "Safe exact matches can be applied automatically after review."
            )

            sales_verification_file = st.file_uploader(
                "Upload Sales File For Verification",
                type=["xlsx"],
                key="sales_verification_file"
            )

            if sales_verification_file is not None:
                sales_uploaded_df, sales_read_error = read_uploaded_excel(
                    sales_verification_file,
                    "Sales Verification file"
                )

                if sales_read_error:
                    st.error(sales_read_error)

                else:
                    verification_result = verify_sales_billing_against_tracker(
                        sales_uploaded_df,
                        tracker
                    )

                    st.markdown("#### Verification Summary")
                    st.dataframe(
                        verification_result["summary"],
                        use_container_width=True
                    )

                    if not verification_result["exceptions"].empty:
                        st.markdown("#### Exceptions To Review")
                        st.warning(
                            f"{len(verification_result['exceptions'])} tracker rows need review before automation."
                        )
                        st.dataframe(
                            verification_result["exceptions"],
                            use_container_width=True
                        )
                        st.download_button(
                            "Download Sales Verification Exceptions",
                            data=to_excel(verification_result["exceptions"]),
                            file_name="sales_billing_verification_exceptions.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                    else:
                        st.success("No tracker billing exceptions found.")

                    if not verification_result["updateable_rows"].empty:
                        st.markdown("#### Rows Safe For Auto Update")
                        st.success(
                            f"{len(verification_result['updateable_rows'])} rows can be updated from sales data."
                        )
                        st.dataframe(
                            verification_result["updateable_rows"],
                            use_container_width=True
                        )
                        st.download_button(
                            "Download Rows Safe For Auto Update",
                            data=to_excel(verification_result["updateable_rows"]),
                            file_name="sales_billing_rows_safe_for_update.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                        confirm_sales_update = st.checkbox(
                            "I reviewed the sales verification and want to update safe tracker rows",
                            key="confirm_sales_billing_auto_update"
                        )

                        if st.button("Apply Sales Billing Auto Update"):
                            if not confirm_sales_update:
                                st.error("Please confirm before applying sales billing updates.")

                            else:
                                sales_update_result = apply_sales_billing_update_upload(
                                    sales_uploaded_df,
                                    tracker
                                )

                                if sales_update_result["updated_rows"] > 0:
                                    st.session_state.latest_billing_update_ids = sales_update_result["updated_ids"]
                                    log_activity(
                                        "sales_billing_upload",
                                        f"{sales_update_result['updated_rows']} tracker rows updated from sales file"
                                    )
                                    st.success(
                                        f"{sales_update_result['updated_rows']} tracker rows updated from sales file"
                                    )
                                    st.rerun()

                                else:
                                    st.warning("No tracker rows were updated from this sales file.")

                    else:
                        st.info("No rows are currently safe for automatic sales update.")

                    if not verification_result["unmatched_sales"].empty:
                        st.markdown("#### Sales Rows Not Found In Tracker")
                        st.info(
                            f"{len(verification_result['unmatched_sales'])} sales billing keys did not match tracker."
                        )
                        st.dataframe(
                            verification_result["unmatched_sales"],
                            use_container_width=True
                        )
                        st.download_button(
                            "Download Unmatched Sales Rows",
                            data=to_excel(verification_result["unmatched_sales"]),
                            file_name="sales_rows_not_found_in_tracker.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                    st.markdown("#### Matched Rows")
                    if verification_result["matches"].empty:
                        st.info("No exact matched billing rows found.")
                    else:
                        st.dataframe(
                            verification_result["matches"],
                            use_container_width=True
                        )
                        st.download_button(
                            "Download Matched Sales Verification Rows",
                            data=to_excel(verification_result["matches"]),
                            file_name="sales_billing_verification_matches.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

        if tracker_task == "Full Tracker Correction Upload":
            render_full_tracker_correction_upload(tracker)

        if tracker_task == "Emergency Manual Allocation Upload":
            st.markdown("---")
            st.subheader("Emergency Manual Allocation Upload")
            st.caption(
                "Use this only when the allocation must be inserted manually because SO/internal data changed."
            )

            manual_allocation_template = get_manual_allocation_template_df()

            with st.expander("Manual allocation upload format"):
                st.dataframe(manual_allocation_template, use_container_width=True)
                st.download_button(
                    "Download Manual Allocation Template",
                    data=to_excel(manual_allocation_template),
                    file_name="manual_allocation_upload_template.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            manual_allocation_file = st.file_uploader(
                "Upload Manual Allocation File",
                type=["xlsx"],
                key="manual_allocation_file"
            )

            if manual_allocation_file is not None:
                manual_uploaded_df, manual_read_error = read_uploaded_excel(
                    manual_allocation_file,
                    "Manual Allocation file"
                )

                if manual_read_error:
                    st.error(manual_read_error)

                else:
                    manual_df, missing_manual_cols = prepare_manual_allocation_upload(
                        manual_uploaded_df
                    )

                    if missing_manual_cols:
                        st.error(f"Missing columns in Manual Allocation file: {missing_manual_cols}")

                    elif manual_df.empty:
                        st.warning("No valid manual allocation rows found.")

                    else:
                        st.dataframe(manual_df, use_container_width=True)

                        if st.button("Save Manual Allocation Upload"):
                            saved_manual_rows = save_manual_allocation_upload(manual_df)
                            log_activity("manual_allocation_upload", f"{saved_manual_rows} manual rows saved")
                            st.success(f"{saved_manual_rows} manual allocation rows saved successfully")
                            st.rerun()

        if tracker_task == "Delete Allocation Upload":
            st.markdown("---")
            st.subheader("Delete Allocation Upload")
            st.caption(
                "Use this when an SO/internal allocation must be revised. Deleted rows are removed from stock blocking."
            )

            delete_template = pd.DataFrame({
                "Tracker ID": [int(tracker.iloc[0]["id"])],
                "Delete Reason": ["SO revised / wrong allocation"],
            })

            with st.expander("Delete upload format"):
                st.dataframe(delete_template, use_container_width=True)
                st.download_button(
                    "Download Delete Allocation Template",
                    data=to_excel(delete_template),
                    file_name="delete_allocation_upload_template.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            delete_file = st.file_uploader(
                "Upload Delete Allocation File",
                type=["xlsx"],
                key="delete_allocation_file"
            )

            if delete_file is not None:
                delete_uploaded_df, delete_read_error = read_uploaded_excel(
                    delete_file,
                    "Delete Allocation file"
                )

                if delete_read_error:
                    st.error(delete_read_error)

                else:
                    st.dataframe(delete_uploaded_df.head(25), use_container_width=True)

                    if st.button("Apply Delete Allocation Upload"):
                        deleted_count, delete_errors = apply_delete_allocation_upload(delete_uploaded_df)

                        if not delete_errors.empty:
                            st.warning(f"{len(delete_errors)} rows could not be deleted.")
                            st.dataframe(delete_errors, use_container_width=True)

                        if deleted_count > 0:
                            st.success(f"{deleted_count} allocation rows deleted successfully")
                            st.rerun()

        if tracker_task == "View / Manual Edit Table":
            if len(tracker) > 500:
                show_all_tracker_rows = st.checkbox(
                    "Show all tracker rows in editable table (slower)",
                    value=False
                )

                editable_source_df = tracker if show_all_tracker_rows else tracker.head(500)
                st.caption(
                    f"Editable table showing {len(editable_source_df):,} of {len(tracker):,} rows."
                )

            else:
                editable_source_df = tracker

            editable_tracker = editable_source_df.copy()

            editable_tracker["sent_for_billing"] = editable_tracker["sent_for_billing"].fillna("No")
            editable_tracker["billing_done"] = editable_tracker["billing_done"].fillna("No")
            editable_tracker["invoice_no"] = editable_tracker["invoice_no"].fillna("")
            editable_tracker["remark"] = editable_tracker["remark"].fillna("")

            if "billed_qty" in editable_tracker.columns:
                editable_tracker["billed_qty"] = editable_tracker["billed_qty"].fillna(0)

            if "balance_to_bill" in editable_tracker.columns:
                editable_tracker["balance_to_bill"] = (
                    editable_tracker["allocated_qty"].fillna(0) -
                    editable_tracker.get("billed_qty", 0)
                ).apply(lambda x: max(clean_number(x), 0))

            editable_tracker["sent_for_billing_tick"] = editable_tracker["sent_for_billing"].apply(
                lambda x: True if str(x).strip().lower() == "yes" else False
            )

            editable_tracker["billing_done_tick"] = editable_tracker["billing_done"].apply(
                lambda x: True if str(x).strip().lower() == "yes" else False
            )

            editable_tracker["sent_date"] = pd.to_datetime(
                editable_tracker["sent_date"],
                errors="coerce"
            ).dt.date

            editable_tracker["billing_date"] = pd.to_datetime(
                editable_tracker["billing_date"],
                errors="coerce"
            ).dt.date

            editable_tracker = editable_tracker.drop(
                columns=["sent_for_billing", "billing_done"],
                errors="ignore"
            )

            tracker_display_columns = [
                "id",
                "allocation_date",
                "po_no",
                "order_id",
                "buyer_code",
                "fsn",
                "title",
                "rr_warehouse",
                "fk_warehouse",
                "sap_code",
                "fcn",
                "appointment_no",
                "appointment_date",
                "pending_amount",
                "allocated_qty",
                "billed_qty",
                "balance_to_bill",
                "sent_date",
                "billing_date",
                "invoice_no",
                "remark",
                "sent_for_billing_tick",
                "billing_done_tick",
            ]

            tracker_display_columns = [
                col for col in tracker_display_columns if col in editable_tracker.columns
            ]

            editable_tracker = editable_tracker[tracker_display_columns]

            st.markdown("---")
            st.subheader("Bulk Update")

            select_all = st.checkbox("Select All Allocation Rows")

            if select_all:
                selected_ids = editable_tracker["id"].tolist()
                st.info(f"{len(selected_ids)} rows selected.")
            else:
                selected_ids = st.multiselect(
                    "Select Allocation IDs",
                    options=editable_tracker["id"].tolist()
                )

            b1, b2, b3, b4 = st.columns(4)

            with b1:
                bulk_sent_tick = st.checkbox("Mark Sent for Billing")

            with b2:
                bulk_sent_date = st.date_input("Sent Date")

            with b3:
                bulk_billing_tick = st.checkbox("Mark Billing Done")

            with b4:
                bulk_billing_date = st.date_input("Billing Date")

            b5, b6 = st.columns(2)

            with b5:
                bulk_invoice_no = st.text_input("Invoice No. Optional")

            with b6:
                bulk_remark = st.text_input("Remark Optional")

            if st.button("Apply Bulk Update"):
                if len(selected_ids) == 0:
                    st.error("Please select at least one allocation row.")

                elif not bulk_sent_tick and not bulk_billing_tick and bulk_invoice_no == "" and bulk_remark == "":
                    st.error("Please choose at least one update action.")

                else:
                    for allocation_id in selected_ids:
                        if bulk_sent_tick:
                            db_execute("""
                            UPDATE allocation_tracker
                            SET
                                sent_for_billing = :sent_for_billing,
                                sent_date = :sent_date
                            WHERE id = :id
                            """, {
                                "sent_for_billing": "Yes",
                                "sent_date": str(bulk_sent_date),
                                "id": int(allocation_id)
                            }, clear_cache=False)

                        if bulk_billing_tick:
                            if "billed_qty" in tracker.columns and "balance_to_bill" in tracker.columns:
                                db_execute("""
                                UPDATE allocation_tracker
                                SET
                                    billing_done = :billing_done,
                                    billing_date = :billing_date,
                                    billed_qty = allocated_qty,
                                    balance_to_bill = 0
                                WHERE id = :id
                                """, {
                                    "billing_done": "Yes",
                                    "billing_date": str(bulk_billing_date),
                                    "id": int(allocation_id)
                                }, clear_cache=False)

                            else:
                                db_execute("""
                                UPDATE allocation_tracker
                                SET
                                    billing_done = :billing_done,
                                    billing_date = :billing_date
                                WHERE id = :id
                                """, {
                                    "billing_done": "Yes",
                                    "billing_date": str(bulk_billing_date),
                                    "id": int(allocation_id)
                                }, clear_cache=False)

                        if bulk_invoice_no.strip() != "":
                            db_execute("""
                            UPDATE allocation_tracker
                            SET invoice_no = :invoice_no
                            WHERE id = :id
                            """, {
                                "invoice_no": bulk_invoice_no,
                                "id": int(allocation_id)
                            }, clear_cache=False)

                        if bulk_remark.strip() != "":
                            db_execute("""
                            UPDATE allocation_tracker
                            SET remark = :remark
                            WHERE id = :id
                            """, {
                                "remark": bulk_remark,
                                "id": int(allocation_id)
                            }, clear_cache=False)

                    st.cache_data.clear()
                    st.success("Bulk update applied successfully")
                    st.rerun()

            st.markdown("---")

            edited_df = st.data_editor(
                editable_tracker,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "id": st.column_config.NumberColumn("ID", disabled=True),
                    "allocation_date": st.column_config.TextColumn("Allocation Date", disabled=True),
                    "po_no": st.column_config.TextColumn("PO No.", disabled=True),
                    "order_id": st.column_config.TextColumn("Order ID", disabled=True),
                    "buyer_code": st.column_config.TextColumn("Buyer Code", disabled=True),
                    "fsn": st.column_config.TextColumn("FSN", disabled=True),
                    "title": st.column_config.TextColumn("Title", disabled=True),
                    "rr_warehouse": st.column_config.TextColumn("RR Warehouse", disabled=True),
                    "fk_warehouse": st.column_config.TextColumn("FK Warehouse", disabled=True),
                    "sap_code": st.column_config.TextColumn("SAP Code", disabled=True),
                    "fcn": st.column_config.TextColumn("FCN", disabled=True),
                    "appointment_no": st.column_config.TextColumn("Appointment No.", disabled=True),
                    "appointment_date": st.column_config.TextColumn("Appointment Date", disabled=True),
                    "pending_amount": st.column_config.NumberColumn("Pending Amount", disabled=True),
                    "allocated_qty": st.column_config.NumberColumn("Allocated Qty.", disabled=True),
                    "billed_qty": st.column_config.NumberColumn("Billed Qty."),
                    "balance_to_bill": st.column_config.NumberColumn("Balance To Bill", disabled=True),
                    "sent_date": st.column_config.DateColumn("Sent Date", format="DD-MM-YYYY"),
                    "billing_date": st.column_config.DateColumn("Billing Date", format="DD-MM-YYYY"),
                    "invoice_no": st.column_config.TextColumn("Invoice No."),
                    "remark": st.column_config.TextColumn("Remark"),
                    "sent_for_billing_tick": st.column_config.CheckboxColumn("Sent for Billing?"),
                    "billing_done_tick": st.column_config.CheckboxColumn("Billing Done?")
                }
            )

            if st.button("Save Manual Table Updates"):
                for _, row in edited_df.iterrows():
                    sent_value = "Yes" if row["sent_for_billing_tick"] else "No"
                    billed_value = "Yes" if row["billing_done_tick"] else "No"
                    manual_billed_qty = clean_number(row.get("billed_qty", 0))
                    manual_allocated_qty = clean_number(row.get("allocated_qty", 0))
                    manual_balance_to_bill = max(manual_allocated_qty - manual_billed_qty, 0)

                    if "billed_qty" in edited_df.columns:
                        billed_value = "Yes" if manual_balance_to_bill <= 0 and manual_allocated_qty > 0 else "No"

                    db_execute("""
                    UPDATE allocation_tracker
                    SET
                        sent_for_billing = :sent_for_billing,
                        sent_date = :sent_date,
                        billing_done = :billing_done,
                        billing_date = :billing_date,
                        invoice_no = :invoice_no,
                        remark = :remark
                    WHERE id = :id
                    """, {
                        "sent_for_billing": sent_value,
                        "sent_date": str(row["sent_date"]) if pd.notna(row["sent_date"]) else "",
                        "billing_done": billed_value,
                        "billing_date": str(row["billing_date"]) if pd.notna(row["billing_date"]) else "",
                        "invoice_no": row["invoice_no"],
                        "remark": row["remark"],
                        "id": int(row["id"])
                    }, clear_cache=False)

                    if "billed_qty" in edited_df.columns and "balance_to_bill" in tracker.columns:
                        db_execute("""
                        UPDATE allocation_tracker
                        SET
                            billed_qty = :billed_qty,
                            balance_to_bill = :balance_to_bill
                        WHERE id = :id
                        """, {
                            "billed_qty": manual_billed_qty,
                            "balance_to_bill": manual_balance_to_bill,
                            "id": int(row["id"])
                        }, clear_cache=False)

                st.cache_data.clear()
                st.success("Manual tracker updates saved successfully")
                st.rerun()


# =====================================================
# BILLING SUMMARY
# =====================================================

elif menu == "Billing Summary":
    render_page_header(
        "Billing Summary",
        "Generate FSN barcode Excel and sticker PDFs from tracker data or urgent direct uploads."
    )

    fsn_barcode_cache = {}
    st.info("Barcode images are generated automatically from FSN. No barcode image upload is required.")

    st.markdown("### Urgent Direct Barcode Download")
    st.caption(
        "Use this when billing needs to happen quickly without updating allocation tracker rows first."
    )

    direct_sample = pd.DataFrame({
        "Invoice No.": ["INV001"],
        "PO No.": ["PO1001"],
        "FSN": ["FSN001"],
        "Title": ["Ceiling Fan 1200mm"],
        "RR Warehouse": ["WH1"],
        "FK Warehouse": ["FK-BLR"],
        "SAP Code": ["SAP-A1"],
        "Qty.": [10]
    })

    with st.expander("Direct billing upload format"):
        st.dataframe(direct_sample, use_container_width=True)
        st.download_button(
            "Download Direct Billing Template",
            data=to_excel(direct_sample),
            file_name="direct_billing_upload_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    direct_billing_file = st.file_uploader(
        "Upload Direct Billing Details File",
        type=["xlsx"],
        key="direct_billing_file"
    )

    if direct_billing_file is not None:
        try:
            direct_uploaded_df = pd.read_excel(direct_billing_file)
        except Exception as e:
            direct_uploaded_df = None
            st.error(
                "Could not read the Direct Billing Details file. "
                "Please upload a normal Excel workbook with at least one visible sheet. "
                f"Original error: {e}"
            )

        if direct_uploaded_df is not None:
            direct_billing_df, missing_direct_cols = prepare_direct_billing_df(
                direct_uploaded_df
            )

            if missing_direct_cols:
                st.error(f"Missing columns in Direct Billing file: {missing_direct_cols}")

            elif direct_billing_df.empty:
                st.warning("No valid FSN rows found in the direct billing file.")

            else:
                d1, d2, d3 = st.columns(3)
                d1.metric("Rows Uploaded", len(direct_billing_df))
                d2.metric("Total Qty", f"{direct_billing_df['Qty.'].sum():,.0f}")
                d3.metric("Generated Barcodes", direct_billing_df["FSN"].nunique())

                st.dataframe(direct_billing_df, use_container_width=True)

                q1, q2 = st.columns(2)

                with q1:
                    st.download_button(
                        "Download Direct Barcode Excel",
                        data=to_excel_billing_with_exact_barcodes(direct_billing_df),
                        file_name="direct_billing_barcode_sheet.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                with q2:
                    st.download_button(
                        "Download Direct Sticker PDF",
                        data=generate_sticker_pdf(direct_billing_df),
                        file_name="direct_warehouse_sticker_sheet.pdf",
                        mime="application/pdf"
                    )

    st.markdown("---")
    st.markdown("### Tracker Billing Workflow")

    latest_billing_update_ids = st.session_state.get("latest_billing_update_ids", [])

    if latest_billing_update_ids:
        billing_scope = st.radio(
            "Billing Summary Scope",
            ["Latest Billing Update Upload", "All Billed Tracker Rows"],
            horizontal=True
        )

    else:
        billing_scope = "All Billed Tracker Rows"

    if billing_scope == "Latest Billing Update Upload":
        billing_df = get_billing_summary(latest_billing_update_ids)
        st.caption(
            f"Showing rows from the latest billing update upload only "
            f"({len(latest_billing_update_ids)} tracker rows updated)."
        )

    else:
        billing_df = get_billing_summary()

    if billing_df.empty:
        if billing_scope == "Latest Billing Update Upload":
            st.info(
                "No invoice rows found in the latest billing update upload. "
                "Check that Invoice No., Billing Date, and Billing Done are filled."
            )
        else:
            st.info(
                "No rows found. Mark 'Sent for Billing' as Yes and enter Invoice No. in Allocation Tracker."
            )

    else:
        billing_df = billing_df.rename(columns={
            "id": "Tracker ID",
            "invoice_no": "Invoice No.",
            "po_no": "PO No.",
            "fsn": "FSN",
            "title": "Title",
            "rr_warehouse": "RR Warehouse",
            "fk_warehouse": "FK Warehouse",
            "sap_code": "SAP Code",
            "allocated_qty": "Qty."
        })

        st.markdown("### Invoice-wise Billing Lines")

        st.dataframe(
            billing_df,
            use_container_width=True
        )

        st.markdown("---")
        st.markdown("### Invoice-wise Billing Lines With Barcode")

        for _, row in billing_df.iterrows():
            fsn = clean_text(row["FSN"])

            with st.container():
                c1, c2, c3, c4, c5, c6 = st.columns(
                    [2, 2, 3, 4, 2, 2]
                )

                with c1:
                    st.markdown("**Invoice No.**")
                    st.write(row["Invoice No."])

                with c2:
                    st.markdown("**PO No.**")
                    st.write(row["PO No."])

                with c3:
                    st.markdown("**FSN**")
                    st.write(row["FSN"])

                with c4:
                    st.markdown("**FSN Barcode**")

                    barcode_bytes = get_fsn_barcode_bytes(fsn, fsn_barcode_cache)

                    if barcode_bytes:
                        st.image(
                            barcode_bytes,
                            width=300
                        )
                    else:
                        st.warning("Barcode Not Found")

                with c5:
                    st.markdown("**SAP Code**")
                    st.write(row["SAP Code"])

                with c6:
                    st.markdown("**Qty.**")
                    st.write(row["Qty."])

                st.markdown("---")

        invoice_summary = billing_df.groupby(
            "Invoice No."
        )["Qty."].sum().reset_index()

        invoice_summary = invoice_summary.rename(
            columns={"Qty.": "Total Qty."}
        )

        st.markdown("### Invoice-wise Total Qty")

        st.dataframe(
            invoice_summary,
            use_container_width=True
        )

        st.download_button(
            "Download Billing Summary With Exact Barcode",
            data=to_excel_billing_with_exact_barcodes(billing_df),
            file_name="billing_summary_with_exact_barcode.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.download_button(
            "Download Warehouse Sticker Sheet PDF",
            data=generate_sticker_pdf(billing_df),
            file_name="warehouse_sticker_sheet.pdf",
            mime="application/pdf"
        )


# =====================================================
# USER MANAGEMENT
# =====================================================

elif menu == "User Management":
    render_page_header(
        "User Management",
        "Create team logins, change roles, and manage account access."
    )

    if st.session_state.role != "Admin":
        st.error("Only Admin users can access this screen.")
        st.stop()

    st.markdown("### Create User")

    u1, u2, u3 = st.columns(3)

    with u1:
        new_username = st.text_input("New Username")

    with u2:
        new_password = st.text_input("New Password", type="password")

    with u3:
        new_role = st.selectbox(
            "Role",
            ["Ops", "Billing", "Viewer", "Admin"]
        )

    if st.button("Create User"):
        new_username = clean_text(new_username)

        if new_username == "":
            st.error("Please enter a username.")

        elif len(new_password) < 6:
            st.error("Password should be at least 6 characters.")

        else:
            try:
                db_execute("""
                    INSERT INTO app_users (username, password_hash, role, active)
                    VALUES (:username, :password_hash, :role, TRUE)
                """, {
                    "username": new_username,
                    "password_hash": hash_password(new_password),
                    "role": new_role
                })

                log_activity(
                    "create_user",
                    f"Created user {new_username} with role {new_role}"
                )
                st.success(f"User {new_username} created successfully.")
                st.rerun()

            except Exception as e:
                st.error(f"Could not create user. It may already exist. Error: {e}")

    st.markdown("---")
    st.markdown("### Manage Existing Users")

    users_df = db_read("""
        SELECT
            id,
            username,
            role,
            active,
            created_at
        FROM app_users
        ORDER BY id
    """)

    st.markdown("#### Change User Access")

    selected_username = st.selectbox(
        "Select User",
        options=users_df["username"].tolist()
    )

    selected_user_row = users_df[users_df["username"] == selected_username].iloc[0]

    a1, a2 = st.columns(2)

    with a1:
        selected_role = st.selectbox(
            "New Role",
            ["Admin", "Ops", "Billing", "Viewer"],
            index=["Admin", "Ops", "Billing", "Viewer"].index(selected_user_row["role"])
        )

    with a2:
        selected_active = st.checkbox(
            "Active User",
            value=bool(selected_user_row["active"])
        )

    if st.button("Update Selected User"):
        would_remove_self_admin = (
            selected_username == st.session_state.username and
            (selected_role != "Admin" or not selected_active)
        )

        other_active_admins = users_df[
            (users_df["username"] != selected_username) &
            (users_df["role"] == "Admin") &
            (users_df["active"] == True)
        ]

        if would_remove_self_admin:
            st.error("You cannot remove your own active Admin access while logged in.")

        elif selected_role != "Admin" and other_active_admins.empty:
            st.error("At least one active Admin user is required.")

        elif not selected_active and selected_role == "Admin" and other_active_admins.empty:
            st.error("At least one active Admin user is required.")

        else:
            db_execute("""
                UPDATE app_users
                SET
                    role = :role,
                    active = :active
                WHERE id = :id
            """, {
                "role": selected_role,
                "active": selected_active,
                "id": int(selected_user_row["id"])
            })

            log_activity(
                "update_user",
                f"Updated {selected_username} to role {selected_role}, active={selected_active}"
            )
            st.success(f"{selected_username} updated successfully.")
            st.rerun()

    st.markdown("#### User Table")

    edited_users_df = st.data_editor(
        users_df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "id": st.column_config.NumberColumn("ID", disabled=True),
            "username": st.column_config.TextColumn("Username", disabled=True),
            "role": st.column_config.SelectboxColumn(
                "Role",
                options=["Admin", "Ops", "Billing", "Viewer"],
                required=True
            ),
            "active": st.column_config.CheckboxColumn("Active"),
            "created_at": st.column_config.TextColumn("Created At", disabled=True),
        }
    )

    if st.button("Save User Changes"):
        active_admin_count = len(
            edited_users_df[
                (edited_users_df["role"] == "Admin") &
                (edited_users_df["active"] == True)
            ]
        )

        current_user_row = edited_users_df[
            edited_users_df["username"] == st.session_state.username
        ]

        if active_admin_count == 0:
            st.error("At least one active Admin user is required.")

        elif (
            not current_user_row.empty and
            (
                current_user_row.iloc[0]["role"] != "Admin" or
                not bool(current_user_row.iloc[0]["active"])
            )
        ):
            st.error("You cannot remove your own active Admin access while logged in.")

        else:
            for _, user_row in edited_users_df.iterrows():
                db_execute("""
                    UPDATE app_users
                    SET
                        role = :role,
                        active = :active
                    WHERE id = :id
                """, {
                    "role": user_row["role"],
                    "active": bool(user_row["active"]),
                    "id": int(user_row["id"])
                })

            log_activity("update_users", "Updated user roles or active status")
            st.success("User changes saved successfully.")
            st.rerun()


# =====================================================
# OPEN ALLOCATION
# =====================================================

elif menu == "Open Allocation Qty":
    render_page_header(
        "Open Allocation Qty",
        "Stock already sent for billing but not yet marked as billed."
    )

    open_alloc = get_open_allocation_qty()

    if open_alloc.empty:
        st.info("No open allocations found.")

    else:
        st.dataframe(
            open_alloc,
            use_container_width=True
        )
